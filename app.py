from __future__ import annotations

import tempfile
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ============================================================
# 每天只需要检查这里
# ============================================================
from pathlib import Path

# 固定分析年份
ANALYSIS_YEAR = 2026

# 月份配置：以后切换月份时优先改这里
REPORT_MONTH = 8
REPORT_MONTH_LABEL = f"{REPORT_MONTH}月"
CUMULATIVE_MONTHS = [7, 8]
CUMULATIVE_MONTH_LABEL = "-".join(str(month) for month in CUMULATIVE_MONTHS) + "月"


# 下面这些口径一般不需要每天修改。
SHEET_NAME = "保单汇总列表"
TEAM_KEYWORD = "大湾区计划"

PVI_THRESHOLD = 6000.0
SUNSHINE_COMPANY_KEYWORD = "阳光人寿"
SUNSHINE_THRESHOLD = 100_000.0
MDRT_REGION_TARGET = 5_000_000.0
LOCAL_AGENT_THRESHOLD = 20_000.0

SHEET_NAMES = {
    "tracking": f"{REPORT_MONTH_LABEL}安盛活力星达成追踪",
    "sunshine": "(阳光)七八联动追踪",
    "mdrt": "MDRT",
    "mdrt_region": "MDRT-区域",
    "local_agent": "大湾区属地代理人达成情况",
    "region_month_pvi": f"区域{CUMULATIVE_MONTH_LABEL}PVI汇总",
    "agency_team_pvi": f"机构团队{REPORT_MONTH_LABEL}PVI汇总",
    "agency_team_hires": f"机构团队{REPORT_MONTH_LABEL}入职人数",
    "kai_xuan_detail": "大湾区凯旋明细（全年）",
    "yaokun_detail": "曜坤区明细（全年）",
    "fangyuan_hq_detail": "方圆区总部直辖明细（全年）",
}

TRACKING_SHEET_NAME = SHEET_NAMES["tracking"]
SUNSHINE_SHEET_NAME = SHEET_NAMES["sunshine"]
REGION_MONTH_PVI_SHEET_NAME = SHEET_NAMES["region_month_pvi"]
AGENCY_TEAM_PVI_SHEET_NAME = SHEET_NAMES["agency_team_pvi"]
AGENCY_TEAM_HIRES_SHEET_NAME = SHEET_NAMES["agency_team_hires"]

FINAL_SHEETS = list(SHEET_NAMES.values())

CUMULATIVE_PVI_COLUMN = f"{CUMULATIVE_MONTH_LABEL}合计PVI"
REPORT_MONTH_PVI_COLUMN = f"{REPORT_MONTH_LABEL}总PVI"
REPORT_MONTH_HIRES_COLUMN = f"{REPORT_MONTH_LABEL}入职人数"

MONTH_PVI_COLUMNS = [f"{month}月总PVI" for month in CUMULATIVE_MONTHS]
MONTH_PVI_MAP = dict(zip(CUMULATIVE_MONTHS, MONTH_PVI_COLUMNS))

REQUIRED_COLUMNS = [
    "出单团队",
    "出单机构",
    "出单人所属区域",
    "出单人所属组织",
    "管理所属区域",
    "管理所属组",
    "服务代理人姓名",
    "服务代理人工号",
    "出单代理人工号",
    "出单代理人姓名",
    "保单号",
    "投保单号",
    "PVI",
    "交单日期",
    "承保日期",
    "承保进度",
    "保单状态",
    "险种名称",
    "产品统计分类",
    "保险公司",
    "回执日期",
    "回访完成日期",
]

EXCLUDED_POLICY_STATUSES = {
    "撤销",
    "犹豫期退保",
    "未承保",
    "待承保",
}

MDRT_LONG_TERM_LIFE_CATEGORIES = {
    "定期寿险",
    "两全险",
    "年金险",
    "终身寿险",
    "重疾险",
}

LOCAL_AGENT_ROSTER = [
    {"所属区域": "丰盛区", "所属组织": "伍尚明直辖组", "工号": "AXA004090", "代理人姓名": "王诗婕", "考核截止时间": "2026年8月31日", "入职日期": "2026-05-18"},
    {"所属区域": "SCREM 关爱区", "所属组织": "帅玉芬直辖组", "工号": "AXA004042", "代理人姓名": "余素娴", "考核截止时间": "2026年9月30日", "入职日期": "2026-06-23"},
    {"所属区域": "SCREM 关爱区", "所属组织": "卢海玉直辖组", "工号": "AXA004121", "代理人姓名": "黄丽枝", "考核截止时间": "2026年9月30日", "入职日期": "2026-06-11"},
    {"所属区域": "SCREM 关爱区", "所属组织": "卢海玉直辖组", "工号": "AXA004141", "代理人姓名": "叶伟灿", "考核截止时间": "2026年9月30日", "入职日期": "2026-06-29"},
    {"所属区域": "蓝天区", "所属组织": "朱申萍直辖组", "工号": "AXA004106", "代理人姓名": "段远婷", "考核截止时间": "2026年9月30日", "入职日期": "2026-06-30"},
    {"所属区域": "SCREM 关爱区", "所属组织": "卢海玉直辖组", "工号": "AXA004139", "代理人姓名": "叶小梨", "考核截止时间": "2026年10月31日", "入职日期": "2026-07-02"},
    {"所属区域": "SCREM 关爱区", "所属组织": "卢海玉直辖组", "工号": "AXA004144", "代理人姓名": "黄伟烈", "考核截止时间": "2026年10月31日", "入职日期": "2026-07-02"},
    {"所属区域": "SCREM 关爱区", "所属组织": "卢海玉直辖组", "工号": "AXA004142", "代理人姓名": "叶鹏娜", "考核截止时间": "2026年10月31日", "入职日期": "2026-07-02"},
    {"所属区域": "SCREM 关爱区", "所属组织": "卢海玉直辖组", "工号": "AXA004165", "代理人姓名": "卢惠玲", "考核截止时间": "2026年10月31日", "入职日期": "2026-07-08"},
    {"所属区域": "SCREM 关爱区", "所属组织": "卢海玉直辖组", "工号": "AXA004140", "代理人姓名": "黄后仪", "考核截止时间": "2026年10月31日", "入职日期": "2026-07-15"},
    {"所属区域": "SCREM 关爱区", "所属组织": "卢海玉直辖组", "工号": "AXA004178", "代理人姓名": "陈浩棠", "考核截止时间": "2026年10月31日", "入职日期": "2026-07-15"},
    {"所属区域": "旭日&旭日北辰区", "所属组织": "林辰龙直辖组", "工号": "AXA003839", "代理人姓名": "黄旭娜", "考核截止时间": "2026年10月31日", "入职日期": "2026-07-16"},
    {"所属区域": "SCREM 关爱区", "所属组织": "卢海玉直辖组", "工号": "AXA004205", "代理人姓名": "郑翠雯", "考核截止时间": "2026年10月31日", "入职日期": "2026-07-24"},
    {"所属区域": "SCREM 关爱区", "所属组织": "卢海玉直辖组", "工号": "AXA004158", "代理人姓名": "余思敏", "考核截止时间": "2026年10月31日", "入职日期": "2026-07-24"},
    {"所属区域": "SCREM 关爱区", "所属组织": "卢海玉直辖组", "工号": "AXA004138", "代理人姓名": "孙敬珉", "考核截止时间": "2026年10月31日", "入职日期": "2026-07-24"},
    {"所属区域": "SCREM 关爱区", "所属组织": "卢海玉直辖组", "工号": "AXA004211", "代理人姓名": "陈咏珊", "考核截止时间": "2026年10月31日", "入职日期": "2026-07-27"},
    {"所属区域": "SCREM 关爱区", "所属组织": "卢海玉直辖组", "工号": "AXA004206", "代理人姓名": "王燕仪", "考核截止时间": "2026年10月31日", "入职日期": "2026-07-27"},
    {"所属区域": "SCREM 关爱区", "所属组织": "卢海玉直辖组", "工号": "AXA004184", "代理人姓名": "刘燕婷", "考核截止时间": "2026年10月31日", "入职日期": "2026-07-27"},
    {"所属区域": "旭日&旭日北辰区", "所属组织": "林辰龙直辖组", "工号": "AXA004218", "代理人姓名": "李育萍", "考核截止时间": "2026年11月30日", "入职日期": "2026-08-03"},
    {"所属区域": "旭日&旭日北辰区", "所属组织": "林辰龙直辖组", "工号": "AXA004219", "代理人姓名": "张媄淇", "考核截止时间": "2026年11月30日", "入职日期": "2026-08-04"},
]

def latest_source_file() -> Path:
    selected = SELECTED_INPUT_FILE if "SELECTED_INPUT_FILE" in globals() else find_source_file()
    print(f"本次分析的表格：{selected.name}")
    print(f"表格完整路径：{selected}")
    return selected


def text_series(series: pd.Series) -> pd.Series:
    return series.astype("string").str.strip()


def parse_pvi(series: pd.Series) -> pd.Series:
    cleaned = text_series(series).str.replace(r"[,，￥¥\s]", "", regex=True)
    return pd.to_numeric(cleaned, errors="coerce").fillna(0.0)



def money(value: float) -> str:
    return f"{value:,.2f}".rstrip("0").rstrip(".")

def empty_frame(columns: list[str]) -> pd.DataFrame:
    return pd.DataFrame(columns=columns)


def normalize_region_name(value: object) -> str:
    text = "" if pd.isna(value) else str(value).strip()
    if not text:
        return "未填写"

    text = text.split("-", 1)[0].strip()

    aliases = {
        "Bees 旭日区": "旭日&旭日北辰区",
        "Bees 旭日北辰区": "旭日&旭日北辰区",
    }

    return aliases.get(text, text)



def clean_data(path: Path, sheet: str, team_keyword: str) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name=sheet, dtype=object)
    df.columns = df.columns.astype(str).str.strip()

    missing = [column for column in REQUIRED_COLUMNS if column not in df.columns]
    if missing:
        raise ValueError(f"工作表缺少必要字段：{missing}")

    df = df.copy()
    df["_pvi"] = parse_pvi(df["PVI"])

    # 全局口径修正：陕西分公司统一归入大湾区计划。
    df.loc[
        text_series(df["出单机构"]).eq("陕西分公司"),
        "出单团队",
    ] = team_keyword

    # 全局口径修正：虚拟总监陕西使用管理/服务字段替换出单字段。
    virtual_shaanxi_mask = (
        text_series(df["出单人所属区域"]).eq("内部测试区域")
        & text_series(df["出单代理人姓名"]).eq("虚拟总监陕西")
    )

    replace_columns = {
        "出单人所属区域": "管理所属区域",
        "出单人所属组织": "管理所属组",
        "出单代理人姓名": "服务代理人姓名",
        "出单代理人工号": "服务代理人工号",
    }

    for target, source in replace_columns.items():
        df.loc[virtual_shaanxi_mask, target] = df.loc[virtual_shaanxi_mask, source]

    date_columns = {
        "交单日期": "_submit_date",
        "承保日期": "_insured_date",
        "回执日期": "_receipt_date",
        "回访完成日期": "_visit_date",
    }

    for source, target in date_columns.items():
        df[target] = pd.to_datetime(df[source], errors="coerce").dt.normalize()

    text_columns = {
        "出单团队": "_team",
        "保单状态": "_status",
        "承保进度": "_progress",
        "出单机构": "_agency",
        "出单人所属区域": "_region",
        "出单人所属组织": "_organization",
        "管理所属区域": "_management_region",
        "管理所属组": "_management_organization",
        "出单代理人工号": "_agent_id",
        "出单代理人姓名": "_agent_name",
        "产品统计分类": "_product_category",
    }

    for source, target in text_columns.items():
        df[target] = text_series(df[source])

    valid_status = (
        df["_status"].notna()
        & df["_status"].ne("")
        & ~df["_status"].isin(EXCLUDED_POLICY_STATUSES)
        & ~df["_status"].str.contains("撤销|犹豫期退保", na=False)
    )

    df = df.loc[valid_status].copy()
    df["_is_team"] = df["_team"].eq(team_keyword)

    # 不在清洗阶段提前筛掉其他团队，后续不同 sheet 会各自筛选。
    df["_is_cancelled"] = df["_status"].str.contains("犹豫期退保|撤销", na=False)
    df["_is_valid"] = ~df["_is_cancelled"]

    return df


def determine_report_date(
    df: pd.DataFrame,
    requested: str | None
) -> pd.Timestamp:

    if requested:
        parsed = pd.to_datetime(requested, errors="raise")
        parsed = pd.Timestamp(parsed).normalize()

        if parsed.year != ANALYSIS_YEAR:
            raise ValueError(
                f"报告日期必须属于{ANALYSIS_YEAR}年，"
                f"当前填写的是：{parsed:%Y-%m-%d}"
            )

        return parsed

    latest = df.loc[
        df["_is_team"]
        & df["_insured_date"].dt.year.eq(ANALYSIS_YEAR),
        "_insured_date"
    ].max()

    if pd.isna(latest):
        raise ValueError(
            f"{ANALYSIS_YEAR}年大湾区数据中没有可用的承保日期"
        )

    return pd.Timestamp(latest).normalize()



def calculate_reports(
    df: pd.DataFrame,
    report_date: pd.Timestamp,
    threshold: float,
) -> dict[str, pd.DataFrame | dict[str, float] | pd.Series]:

    report_date = pd.Timestamp(report_date).normalize()

    # 防止误算其他年份
    if report_date.year != ANALYSIS_YEAR:
        raise ValueError(
            f"报告日期必须属于{ANALYSIS_YEAR}年，"
            f"当前日期是：{report_date:%Y-%m-%d}"
        )

    # 已承保业务按承保年份判断
    insured_in_year = (
        df["_insured_date"].dt.year.eq(ANALYSIS_YEAR)
    )

    # 未承保业务按交单年份判断
    pending_submitted_in_year = (
        df["_insured_date"].isna()
        & df["_submit_date"].dt.year.eq(ANALYSIS_YEAR)
    )

    # 本函数再次过滤，确保所有结果只来自2026年
    df_year = df.loc[
        insured_in_year | pending_submitted_in_year
    ].copy()

    month_start = report_date.replace(day=1)
    month_end = month_start + pd.offsets.MonthEnd(0)

    # 本月已承保数据
    in_period = df_year["_insured_date"].between(
        month_start,
        report_date,
        inclusive="both",
    )

    mtd = df_year.loc[
        df_year["_is_team"] & df_year["_is_valid"] & in_period
    ].copy()

    # 2026年大湾区全部相关记录
    team_rows = df_year.loc[
        df_year["_is_team"]
    ].copy()

    # ============================================================
    # 安盛活力星达成追踪
    # ============================================================
    competition_tracking = (
        mtd.groupby(
            [
                "_agency",
                "_region",
                "_organization",
                "_agent_name",
                "_agent_id",
            ],
            dropna=False,
        )["_pvi"]
        .sum()
        .reset_index()
        .rename(
            columns={
                "_agency": "所属机构",
                "_region": "所属区域",
                "_organization": "所属组织",
                "_agent_name": "代理人姓名",
                "_agent_id": "工号",
                "_pvi": "竞赛业绩",
            }
        )
        .sort_values("竞赛业绩", ascending=False)
    )

    # 只保留本月已经出单的人：按代理人汇总后，竞赛业绩必须大于0
    competition_tracking = competition_tracking.loc[
        competition_tracking["竞赛业绩"].gt(0)
    ].copy()

    competition_tracking["_达标差距数值"] = (
        competition_tracking["竞赛业绩"] - threshold
    )
    competition_tracking["达标差距"] = [
        "达标" if value >= 0 else value
        for value in competition_tracking["_达标差距数值"].fillna(-threshold)
    ]
    competition_tracking = competition_tracking.drop(columns=["_达标差距数值"])
    competition_tracking.insert(0, "序号", range(1, len(competition_tracking) + 1))
    competition_tracking = competition_tracking[
        [
            "序号",
            "所属机构",
            "所属区域",
            "所属组织",
            "代理人姓名",
            "工号",
            "竞赛业绩",
            "达标差距",
        ]
    ]

    qualified_region_summary = (
        competition_tracking.loc[
            competition_tracking["达标差距"].eq("达标")
        ]
        .groupby("所属区域", dropna=False)
        .size()
        .rename("达标数")
        .reset_index()
        .rename(columns={"所属区域": "区域"})
        .sort_values("达标数", ascending=False)
    )
    total_qualified = int(qualified_region_summary["达标数"].sum()) if not qualified_region_summary.empty else 0
    qualified_region_summary = pd.concat(
        [
            qualified_region_summary,
            pd.DataFrame([{"区域": "总计", "达标数": total_qualified}]),
        ],
        ignore_index=True,
    )

    # ============================================================
    # (阳光)七八联动追踪：阳光人寿深圳分公司
    # 口径：大湾区计划 + 阳光人寿 + ANALYSIS_YEAR年累计月份承保 + 有效保单状态。
    # 不使用mtd，避免报告日期只覆盖单月时把七八联动数据筛空。
    # ============================================================
    sunshine_rows = df.loc[
        df["_team"].eq(TEAM_KEYWORD)
        & df["_is_valid"]
        & text_series(df["保险公司"]).str.contains(SUNSHINE_COMPANY_KEYWORD, na=False)
        & df["_insured_date"].dt.year.eq(ANALYSIS_YEAR)
        & df["_insured_date"].dt.month.isin(CUMULATIVE_MONTHS)
    ].copy()

    sunshine_tracking = (
        sunshine_rows.groupby(
            [
                "_agency",
                "_region",
                "_organization",
                "_agent_name",
                "_agent_id",
            ],
            dropna=False,
        )
        .agg(
            件数=("保单号", "count"),
            **{CUMULATIVE_PVI_COLUMN: ("_pvi", "sum")},
        )
        .reset_index()
        .rename(
            columns={
                "_agency": "所属机构",
                "_region": "所属区域",
                "_organization": "所属组织",
                "_agent_name": "代理人姓名",
                "_agent_id": "工号",
            }
        )
        .sort_values(CUMULATIVE_PVI_COLUMN, ascending=False)
    )

    sunshine_tracking = sunshine_tracking.loc[
        sunshine_tracking[CUMULATIVE_PVI_COLUMN].gt(0)
    ].copy()
    sunshine_tracking["_差距数值"] = sunshine_tracking[CUMULATIVE_PVI_COLUMN] - SUNSHINE_THRESHOLD
    sunshine_tracking["差距"] = [
        "达标" if value >= 0 else value
        for value in sunshine_tracking["_差距数值"].fillna(-SUNSHINE_THRESHOLD)
    ]
    sunshine_tracking = sunshine_tracking.drop(columns=["_差距数值"])
    sunshine_tracking.insert(0, "序号", range(1, len(sunshine_tracking) + 1))
    sunshine_tracking = sunshine_tracking[
        [
            "序号",
            "所属机构",
            "所属区域",
            "所属组织",
            "代理人姓名",
            "工号",
            "件数",
            CUMULATIVE_PVI_COLUMN,
            "差距",
        ]
    ]

    sunshine_region_summary = (
        sunshine_tracking.groupby("所属区域", dropna=False)
        .agg(
            符合参赛人数=("工号", "count"),
            达标人数=("差距", lambda values: int((values == "达标").sum())),
        )
        .reset_index()
        .rename(columns={"所属区域": "区域"})
    )
    sunshine_region_summary["未达标人数"] = (
        sunshine_region_summary["符合参赛人数"] - sunshine_region_summary["达标人数"]
    )
    sunshine_region_summary = sunshine_region_summary.sort_values(
        ["达标人数", "符合参赛人数"],
        ascending=[False, False],
    )
    sunshine_region_summary.insert(0, "序号", range(1, len(sunshine_region_summary) + 1))
    sunshine_totals = pd.DataFrame(
        [
            {
                "序号": "总计",
                "区域": "",
                "符合参赛人数": int(sunshine_region_summary["符合参赛人数"].sum()) if not sunshine_region_summary.empty else 0,
                "达标人数": int(sunshine_region_summary["达标人数"].sum()) if not sunshine_region_summary.empty else 0,
                "未达标人数": int(sunshine_region_summary["未达标人数"].sum()) if not sunshine_region_summary.empty else 0,
            }
        ]
    )
    sunshine_region_summary = pd.concat(
        [sunshine_region_summary, sunshine_totals],
        ignore_index=True,
    )

    # ============================================================
    # MDRT：全年个人PVI汇总
    # ============================================================

    # MDRT只做这几步筛选：
    # 1. 不再限制出单团队；保单汇总列表中出单人所属区域只要产生PVI就纳入。
    # 2. 保单状态去掉撤销/犹豫期退保/待承保/空白。
    # 3. 承保日期只保留2026年。
    # 4. 保留所有PVI金额，包含0和负数PVI，用于抵扣全年业绩。
    # 5. 按出单代理人工号合并PVI；团队、机构、区域、组织统一取最后一张承保保单对应信息。
    mdrt_valid = df.loc[
        df["_region"].ne("")
        & df["_insured_date"].dt.year.eq(ANALYSIS_YEAR)
        & df["_status"].notna()
        & df["_status"].ne("")
        & ~df["_status"].isin(["撤销", "犹豫期退保", "待承保"])
        & ~df["_status"].str.contains("撤销|犹豫期退保", na=False)
    ].copy()

    mdrt_valid["_mdrt_region"] = mdrt_valid["_region"]
    fangyuan_mask = mdrt_valid["_region"].eq("方圆区")
    mdrt_valid.loc[
        fangyuan_mask & mdrt_valid["_agency"].eq("总部直辖"),
        "_mdrt_region",
    ] = "方圆区总部直辖"
    mdrt_valid.loc[
        fangyuan_mask & mdrt_valid["_agency"].eq("广东分公司"),
        "_mdrt_region",
    ] = "方圆区广东分公司"

    mdrt_valid["_mdrt_organization"] = mdrt_valid["_organization"]
    virtual_shaanxi_mask = mdrt_valid["_agent_name"].eq("虚拟总监陕西")
    mdrt_valid.loc[
        virtual_shaanxi_mask & mdrt_valid["_management_region"].ne(""),
        "_mdrt_region",
    ] = mdrt_valid.loc[
        virtual_shaanxi_mask & mdrt_valid["_management_region"].ne(""),
        "_management_region",
    ]
    mdrt_valid.loc[
        virtual_shaanxi_mask & mdrt_valid["_management_organization"].ne(""),
        "_mdrt_organization",
    ] = mdrt_valid.loc[
        virtual_shaanxi_mask & mdrt_valid["_management_organization"].ne(""),
        "_management_organization",
    ]

    annual_valid = mdrt_valid.copy()

    mdrt_pvi = (
        mdrt_valid.groupby(["_agent_id", "_mdrt_region"], dropna=False)["_pvi"]
        .sum()
        .rename("全年业绩（PVI）")
        .reset_index()
    )

    mdrt_long_term_pvi = (
        mdrt_valid.loc[
            mdrt_valid["_product_category"].isin(MDRT_LONG_TERM_LIFE_CATEGORIES)
        ]
        .groupby(["_agent_id", "_mdrt_region"], dropna=False)["_pvi"]
        .sum()
        .rename("长期人身险PVI")
        .reset_index()
    )

    mdrt_pvi = mdrt_pvi.merge(
        mdrt_long_term_pvi,
        how="left",
        on=["_agent_id", "_mdrt_region"],
    )
    mdrt_pvi["长期人身险PVI"] = mdrt_pvi["长期人身险PVI"].fillna(0.0)
    mdrt_pvi["占比"] = (
        mdrt_pvi["长期人身险PVI"]
        / mdrt_pvi["全年业绩（PVI）"].where(mdrt_pvi["全年业绩（PVI）"].ne(0))
    ).fillna(0.0)

    mdrt_attrs = (
        mdrt_valid.sort_values("_insured_date")
        .drop_duplicates(["_agent_id", "_mdrt_region"], keep="last")
        [
            [
                "_agent_id",
                "_mdrt_region",
                "_team",
                "_agency",
                "_mdrt_organization",
                "_agent_name",
            ]
        ]
    )

    mdrt_tracking = (
        mdrt_pvi.merge(mdrt_attrs, how="left", on=["_agent_id", "_mdrt_region"])
        .rename(
            columns={
                "_team": "所属团队",
                "_agency": "所属机构",
                "_mdrt_region": "所属区域",
                "_mdrt_organization": "所属组织",
                "_agent_name": "代理人姓名",
                "_agent_id": "工号",
            }
        )
    )

    mdrt_tracking.loc[
        mdrt_tracking["所属区域"].eq("SUPA 威信区"),
        "所属机构",
    ] = "总部直辖"

    mdrt_tracking.loc[
        mdrt_tracking["所属区域"].eq("Bees 旭日区"),
        "所属区域",
    ] = "旭日&旭日北辰区"

    mdrt_tracking.loc[
        mdrt_tracking["所属区域"].eq("RH 喜悦区"),
        "所属机构",
    ] = "广东分公司"

    mdrt_tracking = mdrt_tracking.sort_values(
        "全年业绩（PVI）",
        ascending=False,
    )

    mdrt_tracking = mdrt_tracking.loc[
        mdrt_tracking["全年业绩（PVI）"].gt(0)
    ].copy()

    mdrt_tracking.insert(0, "序号", range(1, len(mdrt_tracking) + 1))

    mdrt_tracking = mdrt_tracking[
        [
            "序号",
            "所属团队",
            "所属机构",
            "所属区域",
            "所属组织",
            "代理人姓名",
            "工号",
            "全年业绩（PVI）",
            "长期人身险PVI",
            "占比",
        ]
    ]

    # MDRT输出前再按工号聚合一次，保证同一个工号只出现一行。
    # 团队、机构、区域、组织优先使用代理人信息表中的最终归属。
    if not mdrt_tracking.empty:
        agent_info_path = SELECTED_AGENT_INFO_FILE.expanduser().resolve()
        if not agent_info_path.exists():
            raise FileNotFoundError(f"没有找到代理人信息表：{agent_info_path}")

        mdrt_agent_info = pd.read_excel(agent_info_path, dtype=object)
        mdrt_agent_info.columns = mdrt_agent_info.columns.astype(str).str.strip()

        required_mdrt_agent_columns = ["工号", "所属团队", "所属机构", "所属区域"]
        missing_mdrt_agent_columns = [
            column for column in required_mdrt_agent_columns
            if column not in mdrt_agent_info.columns
        ]
        if missing_mdrt_agent_columns:
            raise ValueError(f"代理人信息表缺少MDRT归属字段：{missing_mdrt_agent_columns}")

        if "所属组织" not in mdrt_agent_info.columns:
            mdrt_agent_info["所属组织"] = ""

        mdrt_agent_info = mdrt_agent_info.copy()
        mdrt_agent_info["_agent_id"] = text_series(mdrt_agent_info["工号"])
        mdrt_agent_info["_agent_info_team"] = text_series(mdrt_agent_info["所属团队"])
        mdrt_agent_info["_agent_info_agency"] = text_series(mdrt_agent_info["所属机构"])
        mdrt_agent_info["_agent_info_region"] = text_series(mdrt_agent_info["所属区域"])
        mdrt_agent_info["_agent_info_organization"] = text_series(mdrt_agent_info["所属组织"])

        mdrt_agent_info_latest = (
            mdrt_agent_info.loc[mdrt_agent_info["_agent_id"].ne("")]
            .drop_duplicates("_agent_id", keep="last")
            [
                [
                    "_agent_id",
                    "_agent_info_team",
                    "_agent_info_agency",
                    "_agent_info_region",
                    "_agent_info_organization",
                ]
            ]
            .rename(columns={"_agent_id": "工号"})
        )

        mdrt_id_fallback = (
            mdrt_tracking.drop(columns=["序号"])
            .groupby("工号", dropna=False)
            .agg(
                所属团队=("所属团队", "last"),
                所属机构=("所属机构", "last"),
                所属区域=("所属区域", "last"),
                所属组织=("所属组织", "last"),
                代理人姓名=("代理人姓名", "last"),
                **{
                    "全年业绩（PVI）": ("全年业绩（PVI）", "sum"),
                    "长期人身险PVI": ("长期人身险PVI", "sum"),
                },
            )
            .reset_index()
        )

        mdrt_tracking = mdrt_id_fallback.merge(
            mdrt_agent_info_latest,
            how="left",
            on="工号",
        )

        for target_column, agent_info_column in [
            ("所属团队", "_agent_info_team"),
            ("所属机构", "_agent_info_agency"),
            ("所属区域", "_agent_info_region"),
            ("所属组织", "_agent_info_organization"),
        ]:
            has_agent_info_value = (
                mdrt_tracking[agent_info_column].notna()
                & mdrt_tracking[agent_info_column].ne("")
            )
            mdrt_tracking.loc[has_agent_info_value, target_column] = mdrt_tracking.loc[
                has_agent_info_value,
                agent_info_column,
            ]

        mdrt_tracking["占比"] = (
            mdrt_tracking["长期人身险PVI"]
            / mdrt_tracking["全年业绩（PVI）"].where(mdrt_tracking["全年业绩（PVI）"].ne(0))
        ).fillna(0.0)

        mdrt_tracking = mdrt_tracking.sort_values("全年业绩（PVI）", ascending=False)
        mdrt_tracking.insert(0, "序号", range(1, len(mdrt_tracking) + 1))
        mdrt_tracking = mdrt_tracking[
            [
                "序号",
                "所属团队",
                "所属机构",
                "所属区域",
                "所属组织",
                "代理人姓名",
                "工号",
                "全年业绩（PVI）",
                "长期人身险PVI",
                "占比",
            ]
        ]

    # ============================================================
    # MDRT-区域：全年区域PVI汇总
    # ============================================================

    # MDRT-区域基于MDRT个人表汇总，并按区域展示口径做指定合并。
    mdrt_region_merge_map = {
        "顶峰区-陈思慈总监": "顶峰区",
        "顶峰区-龙圻溱总监": "顶峰区",
        "Bees 旭日北辰区": "旭日&旭日北辰区",
        "旭日&旭日北辰区-林辰龙总监": "旭日&旭日北辰区",
        "旭日&旭日北辰区-林平总监": "旭日&旭日北辰区",
        "Bees 旭日区": "旭日&旭日北辰区",
    }
    mdrt_region_source = mdrt_tracking.copy()
    mdrt_region_source["_汇总所属区域"] = (
        mdrt_region_source["所属区域"].replace(mdrt_region_merge_map)
    )
    mdrt_region_summary = (
        mdrt_region_source.groupby(["所属团队", "_汇总所属区域"], dropna=False)["全年业绩（PVI）"]
        .sum()
        .rename("年度PVI")
        .reset_index()
        .rename(columns={"_汇总所属区域": "所属区域"})
        .sort_values("年度PVI", ascending=False)
    )

    mdrt_region_summary = mdrt_region_summary.loc[
        mdrt_region_summary["年度PVI"].gt(0)
    ].copy()

    mdrt_region_summary.insert(
        0,
        "序号",
        range(1, len(mdrt_region_summary) + 1),
    )

    mdrt_region_summary = mdrt_region_summary[
        ["序号", "所属团队", "所属区域", "年度PVI"]
    ]

    # ============================================================
    # 大湾区属地代理人达成情况：固定名单
    # ============================================================

    # ============================================================
    # 大湾区属地代理人达成情况：固定名单
    # ============================================================
    local_agent_tracking = pd.DataFrame(LOCAL_AGENT_ROSTER).copy()
    local_agent_tracking["_start_date"] = pd.to_datetime(
        local_agent_tracking["入职日期"],
        errors="coerce",
    ).dt.normalize()
    local_agent_tracking["_end_date"] = pd.to_datetime(
        local_agent_tracking["考核截止时间"].astype(str).str.replace("年", "-").str.replace("月", "-").str.replace("日", ""),
        errors="coerce",
    ).dt.normalize()

    # 大湾区属地代理人按各自考核周期累计：入职日期 <= 承保日期 <= 考核截止时间。
    # 这里不用 mtd，否则只会统计本月已承保，容易漏掉固定名单里已有保单的人。
    local_agent_pvi_rows = []
    for _, roster_row in local_agent_tracking.iterrows():
        agent_id = roster_row["工号"]
        start_date = roster_row["_start_date"]
        end_date = roster_row["_end_date"]
        agent_rows = df.loc[
            df["_agent_id"].eq(agent_id)
            & df["_insured_date"].between(start_date, end_date, inclusive="both")
        ]
        local_agent_pvi_rows.append(
            {
                "工号": agent_id,
                "PVI（元）": agent_rows["_pvi"].sum(),
            }
        )

    local_agent_pvi = pd.DataFrame(local_agent_pvi_rows)
    local_agent_tracking = local_agent_tracking.merge(
        local_agent_pvi,
        how="left",
        on="工号",
    )
    local_agent_tracking["PVI（元）"] = local_agent_tracking["PVI（元）"].fillna(0.0)
    local_agent_tracking["考核差距（元）"] = local_agent_tracking["PVI（元）"] - LOCAL_AGENT_THRESHOLD
    local_agent_tracking.insert(0, "序号", range(1, len(local_agent_tracking) + 1))
    local_agent_tracking = local_agent_tracking[
        [
            "序号",
            "所属区域",
            "所属组织",
            "工号",
            "代理人姓名",
            "PVI（元）",
            "考核差距（元）",
            "考核截止时间",
            "入职日期",
        ]
    ]

    local_agent_region_summary = (
        local_agent_tracking.groupby("所属区域", dropna=False)
        .agg(
            考核人数=("工号", "count"),
            已达标人数=("考核差距（元）", lambda values: int((values >= 0).sum())),
        )
        .reset_index()
        .rename(columns={"所属区域": "区域"})
    )
    local_agent_region_summary["未达标人数"] = (
        local_agent_region_summary["考核人数"] - local_agent_region_summary["已达标人数"]
    )
    local_agent_region_summary = local_agent_region_summary.sort_values(
        ["考核人数", "区域"],
        ascending=[False, True],
    )
    local_agent_region_summary.insert(0, "序号", range(1, len(local_agent_region_summary) + 1))
    local_agent_total = pd.DataFrame(
        [
            {
                "序号": "合计",
                "区域": "",
                "考核人数": int(local_agent_region_summary["考核人数"].sum()) if not local_agent_region_summary.empty else 0,
                "已达标人数": int(local_agent_region_summary["已达标人数"].sum()) if not local_agent_region_summary.empty else 0,
                "未达标人数": int(local_agent_region_summary["未达标人数"].sum()) if not local_agent_region_summary.empty else 0,
            }
        ]
    )
    local_agent_region_summary = pd.concat(
        [local_agent_region_summary, local_agent_total],
        ignore_index=True,
    )

    # ============================================================
    # 各出单团队/出单人所属区域累计月份PVI汇总
    # 不限制出单团队；同一区域如果属于不同团队，拆成多行展示。
    # 月份按承保日期判断。
    # ============================================================
    region_month_rows = df_year.loc[
        df_year["_is_valid"]
        & df_year["_insured_date"].dt.year.eq(ANALYSIS_YEAR)
        & df_year["_insured_date"].dt.month.isin(CUMULATIVE_MONTHS)
    ].copy()

    if region_month_rows.empty:
        region_july_august_pvi = pd.DataFrame(
            columns=["所属团队", "出单人所属区域", *MONTH_PVI_COLUMNS]
        )
    else:
        region_month_rows["_region_month_summary_region"] = region_month_rows["_region"]
        internal_virtual_shaanxi_mask = (
            region_month_rows["_region"].eq("内部测试区域")
            & region_month_rows["_agent_name"].eq("虚拟总监陕西")
            & region_month_rows["_management_region"].ne("")
        )
        region_month_rows.loc[
            internal_virtual_shaanxi_mask,
            "_region_month_summary_region",
        ] = region_month_rows.loc[
            internal_virtual_shaanxi_mask,
            "_management_region",
        ]
        region_month_rows["承保月份"] = (
            region_month_rows["_insured_date"].dt.month.map(MONTH_PVI_MAP)
        )
        region_july_august_pvi = (
            region_month_rows
            .pivot_table(
                index=["_team", "_region_month_summary_region"],
                columns="承保月份",
                values="_pvi",
                aggfunc="sum",
                fill_value=0.0,
            )
            .reset_index()
            .rename(
                columns={
                    "_team": "所属团队",
                    "_region_month_summary_region": "出单人所属区域",
                }
            )
        )
        for month_column in MONTH_PVI_COLUMNS:
            if month_column not in region_july_august_pvi.columns:
                region_july_august_pvi[month_column] = 0.0
        region_sort_columns = [*reversed(MONTH_PVI_COLUMNS), "所属团队", "出单人所属区域"]
        region_july_august_pvi = region_july_august_pvi[
            ["所属团队", "出单人所属区域", *MONTH_PVI_COLUMNS]
        ].sort_values(
            region_sort_columns,
            ascending=[False] * len(MONTH_PVI_COLUMNS) + [True, True],
        )

    # 区域展示口径合并：先按规则统一区域名称，再按团队+区域重新汇总。
    region_month_merge_map = {
        "Bees 旭日北辰区": "旭日 & 旭日北辰区",
        "Bees 旭日区": "旭日 & 旭日北辰区",
        "旭日 & 旭日北辰区–林辰龙总监": "旭日 & 旭日北辰区",
        "旭日 & 旭日北辰区–林平总监": "旭日 & 旭日北辰区",
        "旭日&旭日北辰区-林辰龙总监": "旭日 & 旭日北辰区",
        "旭日&旭日北辰区-林平总监": "旭日 & 旭日北辰区",
        "旭日&旭日北辰区": "旭日 & 旭日北辰区",
        "顶峰区-陈思慈总监": "顶峰区",
        "顶峰区-龙圻溱总监": "顶峰区",
    }
    if not region_july_august_pvi.empty:
        region_july_august_pvi["出单人所属区域"] = (
            region_july_august_pvi["出单人所属区域"].replace(region_month_merge_map)
        )
        region_july_august_pvi = (
            region_july_august_pvi
            .groupby(["所属团队", "出单人所属区域"], dropna=False)[MONTH_PVI_COLUMNS]
            .sum()
            .reset_index()
        )
        region_sort_columns = [*reversed(MONTH_PVI_COLUMNS), "所属团队", "出单人所属区域"]
        region_july_august_pvi = region_july_august_pvi[
            ["所属团队", "出单人所属区域", *MONTH_PVI_COLUMNS]
        ].sort_values(
            region_sort_columns,
            ascending=[False] * len(MONTH_PVI_COLUMNS) + [True, True],
        )

    region_total = {"所属团队": "", "出单人所属区域": "合计"}
    region_total.update(
        {
            month_column: float(region_july_august_pvi[month_column].sum())
            if month_column in region_july_august_pvi.columns else 0.0
            for month_column in MONTH_PVI_COLUMNS
        }
    )
    region_total_row = pd.DataFrame([region_total])
    region_july_august_pvi = pd.concat(
        [region_july_august_pvi, region_total_row],
        ignore_index=True,
    )


    # ============================================================
    # 各出单机构/出单团队月度PVI汇总
    # 保单状态已在 clean_data 中剔除撤销/犹豫期退保/待承保/空白状态。
    # ============================================================
    agency_team_august_rows = df_year.loc[
        df_year["_is_valid"]
        & df_year["_insured_date"].dt.year.eq(ANALYSIS_YEAR)
        & df_year["_insured_date"].dt.month.eq(REPORT_MONTH)
    ].copy()

    if agency_team_august_rows.empty:
        agency_team_august_pvi = pd.DataFrame(
            columns=["出单机构", "出单团队", REPORT_MONTH_PVI_COLUMN]
        )
    else:
        agency_team_august_pvi = (
            agency_team_august_rows
            .groupby(["_agency", "_team"], dropna=False)["_pvi"]
            .sum()
            .reset_index()
            .rename(
                columns={
                    "_agency": "出单机构",
                    "_team": "出单团队",
                    "_pvi": REPORT_MONTH_PVI_COLUMN,
                }
            )
            .sort_values(
                ["出单机构", REPORT_MONTH_PVI_COLUMN, "出单团队"],
                ascending=[True, False, True],
            )
        )

    agency_team_total_row = pd.DataFrame(
        [
            {
                "出单机构": "合计",
                "出单团队": "",
                REPORT_MONTH_PVI_COLUMN: float(agency_team_august_pvi[REPORT_MONTH_PVI_COLUMN].sum()) if REPORT_MONTH_PVI_COLUMN in agency_team_august_pvi.columns else 0.0,
            }
        ]
    )
    agency_team_august_pvi = pd.concat(
        [agency_team_august_pvi, agency_team_total_row],
        ignore_index=True,
    )

    # ============================================================
    # 各出单机构/出单团队月度入职人数汇总
    # 使用代理人信息表，按入职日期统计报告月份入职人数。
    # ============================================================
    agent_info_path = SELECTED_AGENT_INFO_FILE.expanduser().resolve()
    if not agent_info_path.exists():
        raise FileNotFoundError(f"没有找到代理人信息表：{agent_info_path}")

    agent_info = pd.read_excel(agent_info_path, dtype=object)
    agent_info.columns = agent_info.columns.astype(str).str.strip()

    required_agent_columns = ["工号", "所属机构", "所属团队", "所属区域", "入职日期"]
    missing_agent_columns = [
        column for column in required_agent_columns
        if column not in agent_info.columns
    ]
    if missing_agent_columns:
        raise ValueError(f"代理人信息表缺少必要字段：{missing_agent_columns}")

    agent_info = agent_info.copy()
    agent_info["_hire_date"] = pd.to_datetime(agent_info["入职日期"], errors="coerce").dt.normalize()
    agent_info["_agency"] = text_series(agent_info["所属机构"])
    agent_info["_team"] = text_series(agent_info["所属团队"])
    agent_info["_region"] = text_series(agent_info["所属区域"])
    agent_info["_agent_id"] = text_series(agent_info["工号"])

    report_day = pd.Timestamp(report_date).normalize()

    target_hire_scope = (
        agent_info["_team"].eq(TEAM_KEYWORD)
        | agent_info["_region"].eq("PRINCE BARON 曜坤区")
        | agent_info["_region"].eq("大湾区凯旋")
        | (
            agent_info["_region"].eq("方圆区")
            & agent_info["_agency"].eq("总部直辖")
        )
    )
    today_new_hires = agent_info.loc[
        target_hire_scope
        & agent_info["_hire_date"].eq(report_day)
    ]
    month_new_hires = agent_info.loc[
        target_hire_scope
        & agent_info["_hire_date"].dt.year.eq(ANALYSIS_YEAR)
        & agent_info["_hire_date"].dt.month.eq(REPORT_MONTH)
    ]
    agent_hire_summary = {
        "today_new_hires": int(today_new_hires["_agent_id"].nunique()),
        "month_new_hires": int(month_new_hires["_agent_id"].nunique()),
    }

    august_hires = agent_info.loc[
        agent_info["_hire_date"].dt.year.eq(ANALYSIS_YEAR)
        & agent_info["_hire_date"].dt.month.eq(REPORT_MONTH)
    ].copy()

    if august_hires.empty:
        agency_team_august_hires = pd.DataFrame(
            columns=["出单机构", "出单团队", REPORT_MONTH_HIRES_COLUMN]
        )
    else:
        agency_team_august_hires = (
            august_hires
            .groupby(["_agency", "_team"], dropna=False)["_agent_id"]
            .nunique()
            .reset_index()
            .rename(
                columns={
                    "_agency": "出单机构",
                    "_team": "出单团队",
                    "_agent_id": REPORT_MONTH_HIRES_COLUMN,
                }
            )
            .sort_values(
                ["出单机构", REPORT_MONTH_HIRES_COLUMN, "出单团队"],
                ascending=[True, False, True],
            )
        )

    agency_team_hire_total_row = pd.DataFrame(
        [
            {
                "出单机构": "合计",
                "出单团队": "",
                REPORT_MONTH_HIRES_COLUMN: int(agency_team_august_hires[REPORT_MONTH_HIRES_COLUMN].sum()) if REPORT_MONTH_HIRES_COLUMN in agency_team_august_hires.columns else 0,
            }
        ]
    )
    agency_team_august_hires = pd.concat(
        [agency_team_august_hires, agency_team_hire_total_row],
        ignore_index=True,
    )

    # ============================================================
    # 指定区域明细：按原始保单行单独拉出，并在最后追加PVI汇总
    # ============================================================
    detail_columns = [column for column in df_year.columns if not str(column).startswith("_")]

    def build_detail_sheet(rows: pd.DataFrame) -> pd.DataFrame:
        detail = rows.loc[:, detail_columns].copy()
        if "PVI" in detail.columns:
            detail["PVI"] = parse_pvi(detail["PVI"])
            total_pvi = float(detail["PVI"].sum())
        else:
            total_pvi = float(rows["_pvi"].sum())

        total_row = {column: "" for column in detail.columns}
        if len(detail.columns) > 0:
            total_row[detail.columns[0]] = "PVI汇总"
        if "PVI" in detail.columns:
            total_row["PVI"] = total_pvi

        return pd.concat([detail, pd.DataFrame([total_row])], ignore_index=True)

    kai_xuan_detail = build_detail_sheet(
        df_year.loc[
            df_year["_is_valid"]
            & df_year["_region"].eq("大湾区凯旋")
        ].copy()
    )

    yaokun_detail = build_detail_sheet(
        df_year.loc[
            df_year["_is_valid"]
            & df_year["_region"].eq("PRINCE BARON 曜坤区")
        ].copy()
    )

    fangyuan_hq_detail = build_detail_sheet(
        df_year.loc[
            df_year["_is_valid"]
            & df_year["_region"].eq("方圆区")
            & df_year["_agency"].eq("总部直辖")
        ].copy()
    )

    return {
        "competition_tracking": competition_tracking,
        "qualified_region_summary": qualified_region_summary,
        "sunshine_tracking": sunshine_tracking,
        "sunshine_region_summary": sunshine_region_summary,
        "mdrt_tracking": mdrt_tracking,
        "mdrt_region_summary": mdrt_region_summary,
        "local_agent_tracking": local_agent_tracking,
        "local_agent_region_summary": local_agent_region_summary,
        "region_july_august_pvi": region_july_august_pvi,
        "agency_team_august_pvi": agency_team_august_pvi,
        "agency_team_august_hires": agency_team_august_hires,
        "agent_hire_summary": agent_hire_summary,
        "kai_xuan_detail": kai_xuan_detail,
        "yaokun_detail": yaokun_detail,
        "fangyuan_hq_detail": fangyuan_hq_detail,
    }


def style_excel(
    path: Path,
    reports: dict[str, pd.DataFrame | dict[str, int]],
) -> None:
    from openpyxl import load_workbook

    workbook = load_workbook(path)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin_gray = Side(style="thin", color="D9E1F2")

    blue_fill = PatternFill("solid", fgColor="B4C6E7")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    black_side = Side(style="thin", color="000000")
    table_border = Border(
        left=black_side,
        right=black_side,
        top=black_side,
        bottom=black_side,
    )

    def add_title(
        sheet,
        title: str,
        end_column: int,
        fill_color: str = "203864",
        font_color: str = "FFFFFF",
        font_size: int = 14,
        height: int = 24,
    ) -> None:
        sheet.insert_rows(1)
        sheet.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=end_column,
        )

        cell = sheet.cell(1, 1)
        cell.value = title
        cell.fill = PatternFill("solid", fgColor=fill_color)
        cell.font = Font(color=font_color, bold=True, size=font_size)
        cell.alignment = Alignment(horizontal="center", vertical="center")

        sheet.row_dimensions[1].height = height
        sheet.freeze_panes = "A3"
        sheet.auto_filter.ref = f"A2:{get_column_letter(end_column)}{sheet.max_row}"

    def style_table(
        sheet,
        max_col: int,
        header_color: str = "203864",
        highlight_col: int | None = None,
        highlight_value: str = "达标",
        body_fill: str = "FFFFFF",
    ) -> None:
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=max_col):
            is_header = row[0].row == 2
            is_highlight = (
                highlight_col is not None
                and row[highlight_col - 1].value == highlight_value
            )

            for cell in row:
                cell.border = table_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

                if is_header:
                    cell.fill = PatternFill("solid", fgColor=header_color)
                    cell.font = Font(color="FFFFFF", bold=True)
                elif is_highlight:
                    cell.fill = blue_fill
                else:
                    cell.fill = PatternFill("solid", fgColor=body_fill)

    def set_widths(sheet, widths: list[int], start: int = 1) -> None:
        for offset, width in enumerate(widths):
            sheet.column_dimensions[get_column_letter(start + offset)].width = width

    def write_summary_table(
        sheet,
        summary: pd.DataFrame,
        start_col: int,
        start_row: int,
        headers: list[str],
        total_label: str = "总计",
        red_last_col: bool = False,
    ) -> None:
        for offset, header in enumerate(headers):
            cell = sheet.cell(start_row, start_col + offset)
            cell.value = header
            cell.fill = PatternFill(
                "solid",
                fgColor="C00000" if red_last_col and offset == len(headers) - 1 else "4472C4",
            )
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = table_border

        for index, row_data in summary.iterrows():
            excel_row = start_row + index + 1
            is_total = str(row_data[headers[0]]) == total_label

            for offset, header in enumerate(headers):
                cell = sheet.cell(excel_row, start_col + offset)
                cell.value = row_data[header]
                cell.border = table_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

                if is_total:
                    cell.fill = PatternFill(
                        "solid",
                        fgColor="C00000" if red_last_col and offset == len(headers) - 1 else "4472C4",
                    )
                    cell.font = Font(color="FFFFFF", bold=True)
                elif red_last_col and offset == len(headers) - 1:
                    cell.font = Font(color="FF0000")

    # 只保留最终需要输出的 sheet。
    for sheet in list(workbook.worksheets):
        if sheet.title not in FINAL_SHEETS:
            workbook.remove(sheet)

    # 通用基础格式。
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.sheet_view.showGridLines = False

        if sheet.max_row >= 1 and sheet.max_column >= 1:
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(bottom=thin_gray)

            sheet.auto_filter.ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"

        for column_index in range(1, sheet.max_column + 1):
            values = [
                sheet.cell(row, column_index).value
                for row in range(1, min(sheet.max_row, 200) + 1)
            ]
            max_length = max(
                (len(str(value)) for value in values if value is not None),
                default=8,
            )
            sheet.column_dimensions[get_column_letter(column_index)].width = min(
                max(max_length + 2, 10),
                35,
            )

        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if cell.is_date:
                    cell.number_format = "yyyy-mm-dd"
                elif isinstance(cell.value, float):
                    cell.number_format = "#,##0.00"

    # 安盛活力星达成追踪
    tracking = workbook[TRACKING_SHEET_NAME]
    add_title(tracking, TRACKING_SHEET_NAME, end_column=8)
    style_table(tracking, max_col=8, highlight_col=8)
    set_widths(tracking, [8, 14, 28, 20, 14, 14, 14, 12])

    for row in tracking.iter_rows(min_row=2, max_row=tracking.max_row, max_col=8):
        for cell in row:
            if cell.column == 7 and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"
            if cell.column == 8:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0"
                    cell.font = Font(bold=True)
                elif cell.value == "达标":
                    cell.font = Font(bold=True)

    region_summary = reports.get("qualified_region_summary")
    if isinstance(region_summary, pd.DataFrame):
        write_summary_table(
            tracking,
            region_summary,
            start_col=11,
            start_row=2,
            headers=["区域", "达标数"],
            total_label="总计",
        )
        set_widths(tracking, [28, 12], start=11)

    # 阳光联动
    sunshine_sheet = workbook[SUNSHINE_SHEET_NAME]
    add_title(
        sunshine_sheet,
        "【七八联动 燃聚‘滇’峰】竞赛追踪",
        end_column=9,
        font_size=13,
    )
    style_table(sunshine_sheet, max_col=9, highlight_col=9)
    set_widths(sunshine_sheet, [8, 14, 28, 20, 14, 14, 10, 14, 14])

    for row in sunshine_sheet.iter_rows(min_row=2, max_row=sunshine_sheet.max_row, max_col=9):
        for cell in row:
            if cell.column in {7, 8} and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"
            if cell.column == 9:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0"
                    if cell.value < 0:
                        cell.font = Font(color="FF0000", bold=True)
                elif cell.value == "达标":
                    cell.font = Font(bold=True)

    sunshine_summary = reports.get("sunshine_region_summary")
    if isinstance(sunshine_summary, pd.DataFrame):
        write_summary_table(
            sunshine_sheet,
            sunshine_summary,
            start_col=11,
            start_row=2,
            headers=["序号", "区域", "符合参赛人数", "达标人数", "未达标人数"],
            total_label="总计",
            red_last_col=True,
        )
        set_widths(sunshine_sheet, [8, 24, 16, 14, 14], start=11)

    # MDRT
    mdrt_sheet = workbook["MDRT"]
    add_title(
        mdrt_sheet,
        "全年【百万精英 浪漫之旅】--个人通道竞赛差距",
        end_column=10,
    )
    style_table(mdrt_sheet, max_col=10)
    set_widths(mdrt_sheet, [8, 18, 14, 28, 22, 14, 14, 18, 18, 12])

    for row in mdrt_sheet.iter_rows(min_row=2, max_row=mdrt_sheet.max_row, max_col=10):
        is_header = row[0].row == 2
        for cell in row:
            if not is_header and cell.column in {2, 6}:
                cell.font = Font(bold=True)
            if cell.column in {8, 9} and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"
            if cell.column == 10 and isinstance(cell.value, (int, float)):
                cell.number_format = "0.00%"

    # MDRT-区域
    mdrt_region_sheet = workbook["MDRT-区域"]
    add_title(
        mdrt_region_sheet,
        "总监区域业务进度",
        end_column=4,
        font_size=16,
        height=30,
    )
    style_table(mdrt_region_sheet, max_col=4)
    set_widths(mdrt_region_sheet, [10, 18, 28, 18])

    for row in mdrt_region_sheet.iter_rows(min_row=2, max_row=mdrt_region_sheet.max_row, max_col=4):
        for cell in row:
            if cell.column == 4 and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"

    # 大湾区属地代理人达成情况
    local_agent_sheet = workbook["大湾区属地代理人达成情况"]
    add_title(
        local_agent_sheet,
        "大湾区属地代理人达成情况",
        end_column=9,
        fill_color="B4C6E7",
        font_color="000000",
        font_size=16,
    )
    style_table(
        local_agent_sheet,
        max_col=9,
        header_color="4472C4",
        body_fill="D9E6F2",
    )
    set_widths(local_agent_sheet, [8, 22, 20, 16, 14, 14, 16, 16, 14])

    for row in local_agent_sheet.iter_rows(min_row=2, max_row=local_agent_sheet.max_row, max_col=9):
        for cell in row:
            if cell.column in {6, 7} and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"
                if cell.column == 7 and cell.value < 0:
                    cell.font = Font(color="000000")

    local_summary = reports.get("local_agent_region_summary")
    if isinstance(local_summary, pd.DataFrame):
        write_summary_table(
            local_agent_sheet,
            local_summary,
            start_col=11,
            start_row=2,
            headers=["序号", "区域", "考核人数", "已达标人数", "未达标人数"],
            total_label="合计",
            red_last_col=True,
        )
        set_widths(local_agent_sheet, [8, 24, 14, 14, 14], start=11)

    workbook.save(path)


def export_excel(
    reports: dict[str, pd.DataFrame | dict[str, int]],
    output_path: Path,
    report_date: pd.Timestamp,
) -> None:
    rules = pd.DataFrame(
        [
            ["报告日期", report_date.strftime("%Y-%m-%d")],
            ["数据文件", str(INPUT_FILE) if INPUT_FILE else f"自动读取 {DATA_FOLDER}/{FILE_PATTERN} 中最新文件"],
            ["有效保单状态", "剔除撤销、犹豫期退保、未承保/待承保、空白状态"],
            ["安盛活力星", f"出单团队为{TEAM_KEYWORD}，承保月份为{REPORT_MONTH_LABEL}，按工号汇总PVI并判断是否达到{PVI_THRESHOLD:,.0f}"],
            ["阳光联动", f"出单团队为{TEAM_KEYWORD}，保险公司包含阳光人寿，承保月份属于{CUMULATIVE_MONTH_LABEL}，合计PVI达到{SUNSHINE_THRESHOLD:,.0f}为达标"],
            ["MDRT", f"统计{ANALYSIS_YEAR}年所有有效承保PVI，不限制出单团队；MDRT-区域基于MDRT结果汇总"],
            ["区域累计PVI", f"按所属团队和出单人所属区域汇总{CUMULATIVE_MONTH_LABEL}PVI，并按当前维护规则合并部分区域名称"],
            ["代理人信息表", "用于机构团队月度入职人数、今日新增、本月累计新增，以及MDRT人员归属修正"],
        ],
        columns=["规则", "定义"],
    )

    frames = {
        TRACKING_SHEET_NAME: reports["competition_tracking"],
        SUNSHINE_SHEET_NAME: reports["sunshine_tracking"],
        "MDRT": reports["mdrt_tracking"],
        "MDRT-区域": reports["mdrt_region_summary"],
        "大湾区属地代理人达成情况": reports["local_agent_tracking"],
        REGION_MONTH_PVI_SHEET_NAME: reports["region_july_august_pvi"],
        AGENCY_TEAM_PVI_SHEET_NAME: reports["agency_team_august_pvi"],
        AGENCY_TEAM_HIRES_SHEET_NAME: reports["agency_team_august_hires"],
        "大湾区凯旋明细（全年）": reports["kai_xuan_detail"],
        "曜坤区明细（全年）": reports["yaokun_detail"],
        "方圆区总部直辖明细（全年）": reports["fangyuan_hq_detail"],
    }

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, frame in frames.items():
            assert isinstance(frame, pd.DataFrame)
            frame.to_excel(writer, sheet_name=sheet_name, index=False)

    style_excel(output_path, reports)

st.set_page_config(page_title="PVI自动化报表Demo", page_icon="📊", layout="wide")
st.title("保险经销业务 PVI 自动化报表分析工具")
st.caption("上传保单汇总列表和代理人信息表，自动生成新版多 Sheet Excel 报告。")

with st.sidebar:
    st.header("统计配置")
    report_month = st.number_input("报告月份", min_value=1, max_value=12, value=REPORT_MONTH, step=1)
    cumulative_months = st.multiselect(
        "累计统计月份",
        options=list(range(1, 13)),
        default=CUMULATIVE_MONTHS,
    )
    report_date_text = st.text_input("补算日期（可选，格式 YYYY-MM-DD）", value="")
    st.caption("留空时，系统自动使用数据中 2026 年大湾区计划最新承保日期。")

policy_file = st.file_uploader("上传保单汇总列表 Excel", type=["xlsx"])
agent_file = st.file_uploader("上传代理人信息 Excel", type=["xlsx", "xls"])

st.info(
    "当前版本已移除图片看板、月度目标缺口、目标达成率和信泰旧方案；"
    "保留安盛活力星、阳光联动、MDRT、MDRT-区域、区域累计PVI、机构团队汇总、属地代理人和三类全年明细。"
)

if st.button("生成报告", type="primary", disabled=not (policy_file and agent_file)):
    if not cumulative_months:
        st.error("请至少选择一个累计统计月份。")
        st.stop()

    with st.spinner("正在清洗数据并生成 Excel 报告..."):
        with tempfile.TemporaryDirectory() as tmpdir:
            tmpdir_path = Path(tmpdir)
            policy_path = tmpdir_path / policy_file.name
            agent_path = tmpdir_path / agent_file.name
            output_path = tmpdir_path / "PVI经营报告.xlsx"
            policy_path.write_bytes(policy_file.getbuffer())
            agent_path.write_bytes(agent_file.getbuffer())

            globals()["REPORT_MONTH"] = int(report_month)
            globals()["REPORT_MONTH_LABEL"] = f"{int(report_month)}月"
            globals()["CUMULATIVE_MONTHS"] = sorted(int(month) for month in cumulative_months)
            globals()["CUMULATIVE_MONTH_LABEL"] = "-".join(str(month) for month in CUMULATIVE_MONTHS) + "月"
            globals()["TRACKING_SHEET_NAME"] = f"{REPORT_MONTH_LABEL}安盛活力星达成追踪"
            globals()["REGION_MONTH_PVI_SHEET_NAME"] = f"区域{CUMULATIVE_MONTH_LABEL}PVI汇总"
            globals()["AGENCY_TEAM_PVI_SHEET_NAME"] = f"机构团队{REPORT_MONTH_LABEL}PVI汇总"
            globals()["AGENCY_TEAM_HIRES_SHEET_NAME"] = f"机构团队{REPORT_MONTH_LABEL}入职人数"
            globals()["CUMULATIVE_PVI_COLUMN"] = f"{CUMULATIVE_MONTH_LABEL}合计PVI"
            globals()["REPORT_MONTH_PVI_COLUMN"] = f"{REPORT_MONTH_LABEL}总PVI"
            globals()["REPORT_MONTH_HIRES_COLUMN"] = f"{REPORT_MONTH_LABEL}入职人数"
            globals()["MONTH_PVI_COLUMNS"] = [f"{month}月总PVI" for month in CUMULATIVE_MONTHS]
            globals()["MONTH_PVI_MAP"] = dict(zip(CUMULATIVE_MONTHS, MONTH_PVI_COLUMNS))
            globals()["SHEET_NAMES"] = {
                "tracking": TRACKING_SHEET_NAME,
                "sunshine": SUNSHINE_SHEET_NAME,
                "mdrt": "MDRT",
                "mdrt_region": "MDRT-区域",
                "local_agent": "大湾区属地代理人达成情况",
                "region_month_pvi": REGION_MONTH_PVI_SHEET_NAME,
                "agency_team_pvi": AGENCY_TEAM_PVI_SHEET_NAME,
                "agency_team_hires": AGENCY_TEAM_HIRES_SHEET_NAME,
                "kai_xuan_detail": "大湾区凯旋明细（全年）",
                "yaokun_detail": "曜坤区明细（全年）",
                "fangyuan_hq_detail": "方圆区总部直辖明细（全年）",
            }
            globals()["FINAL_SHEETS"] = list(SHEET_NAMES.values())
            globals()["SELECTED_AGENT_INFO_FILE"] = agent_path
            globals()["INPUT_FILE"] = policy_path
            globals()["DATA_FOLDER"] = policy_path.parent
            globals()["FILE_PATTERN"] = "保单汇总列表*.xlsx"

            requested_date = report_date_text.strip() or None
            df = clean_data(policy_path, SHEET_NAME, TEAM_KEYWORD)
            report_date = determine_report_date(df, requested_date)
            reports = calculate_reports(df=df, report_date=report_date, threshold=PVI_THRESHOLD)
            export_excel(reports, output_path, report_date)
            excel_bytes = output_path.read_bytes()

    st.success(f"报告已生成：{report_date:%Y-%m-%d}")

    agent_hire_summary = reports.get("agent_hire_summary", {})
    col1, col2, col3 = st.columns(3)
    col1.metric("今日新增", f"{agent_hire_summary.get('today_new_hires', 0)} 人")
    col2.metric("本月累计新增", f"{agent_hire_summary.get('month_new_hires', 0)} 人")
    col3.metric("输出 Sheet", f"{len(FINAL_SHEETS)} 个")

    st.download_button(
        "下载 Excel 报告",
        data=excel_bytes,
        file_name=f"PVI经营报告_{report_date:%Y%m%d}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.subheader("本次输出 Sheet")
    st.write(FINAL_SHEETS)
else:
    st.warning("请先上传保单汇总列表和代理人信息表。")
