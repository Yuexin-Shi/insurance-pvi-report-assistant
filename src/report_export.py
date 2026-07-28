from __future__ import annotations

import base64
from pathlib import Path

import pandas as pd
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import *
from metrics import money

def style_excel(
    path: Path,
    dashboard_png: Path,
    reports: dict[str, pd.DataFrame | dict[str, float] | pd.Series],
) -> None:
    from openpyxl import load_workbook

    workbook = load_workbook(path)

    def write_embedded_image(image_base64: str, filename: str) -> Path:
        image_path = path.parent / filename
        image_path.write_bytes(base64.b64decode(image_base64))
        return image_path
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    pale_fill = PatternFill("solid", fgColor="DCE6F1")
    thin_gray = Side(style="thin", color="D9E1F2")

    for sheet in list(workbook.worksheets):
        if sheet.title not in FINAL_SHEETS:
            workbook.remove(sheet)

    for sheet in workbook.worksheets:
        header_row = 1
        sheet.freeze_panes = f"A{header_row + 1}"
        sheet.sheet_view.showGridLines = False
        if sheet.max_row >= header_row and sheet.max_column >= 1:
            for cell in sheet[header_row]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(bottom=thin_gray)
            sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(sheet.max_column)}{sheet.max_row}"

        for column_index in range(1, sheet.max_column + 1):
            values = [sheet.cell(row, column_index).value for row in range(1, min(sheet.max_row, 200) + 1)]
            max_length = max((len(str(value)) for value in values if value is not None), default=8)
            sheet.column_dimensions[get_column_letter(column_index)].width = min(max(max_length + 2, 10), 35)

        for row in sheet.iter_rows(min_row=header_row + 1):
            for cell in row:
                if cell.is_date:
                    cell.number_format = "yyyy-mm-dd"
                elif isinstance(cell.value, float):
                    cell.number_format = "#,##0.00"

    tracking = workbook["7月安盛活力星达成追踪"]
    tracking.insert_rows(1)
    tracking.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    title_cell = tracking.cell(1, 1)
    title_cell.value = "7月安盛活力星达成追踪"
    title_cell.fill = PatternFill("solid", fgColor="203864")
    title_cell.font = Font(color="FFFFFF", bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    tracking.row_dimensions[1].height = 24
    tracking.freeze_panes = "A3"
    tracking.auto_filter.ref = f"A2:H{tracking.max_row}"
    blue_fill = PatternFill("solid", fgColor="B4C6E7")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    black_side = Side(style="thin", color="000000")
    table_border = Border(left=black_side, right=black_side, top=black_side, bottom=black_side)
    for row in tracking.iter_rows(min_row=2, max_row=tracking.max_row, max_col=8):
        is_header = row[0].row == 2
        is_qualified = row[7].value == "达标"
        for cell in row:
            cell.border = table_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if is_header:
                cell.fill = PatternFill("solid", fgColor="203864")
                cell.font = Font(color="FFFFFF", bold=True)
            elif is_qualified:
                cell.fill = blue_fill
            else:
                cell.fill = white_fill
            if cell.column == 7 and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"
            if cell.column == 8 and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"
                cell.font = Font(bold=True)
            if cell.column == 8 and cell.value == "达标":
                cell.font = Font(bold=True)
    for column_index, width in enumerate([8, 14, 28, 20, 14, 14, 14, 12], start=1):
        tracking.column_dimensions[get_column_letter(column_index)].width = width

    region_summary = reports.get("qualified_region_summary")
    if isinstance(region_summary, pd.DataFrame):
        start_col = 11
        start_row = 2
        tracking.cell(start_row, start_col).value = "区域"
        tracking.cell(start_row, start_col + 1).value = "达标数"
        for cell in tracking[start_row][start_col - 1:start_col + 1]:
            cell.fill = PatternFill("solid", fgColor="4472C4")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = table_border

        for index, row_data in region_summary.iterrows():
            excel_row = start_row + index + 1
            tracking.cell(excel_row, start_col).value = row_data["区域"]
            tracking.cell(excel_row, start_col + 1).value = int(row_data["达标数"])
            for cell in (tracking.cell(excel_row, start_col), tracking.cell(excel_row, start_col + 1)):
                cell.border = table_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if row_data["区域"] == "总计":
                    cell.font = Font(bold=True)

        tracking.column_dimensions[get_column_letter(start_col)].width = 28
        tracking.column_dimensions[get_column_letter(start_col + 1)].width = 12

    sunshine_sheet = workbook["(阳光)七八联动七月追踪"]
    sunshine_sheet.insert_rows(1)
    sunshine_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    sunshine_title = sunshine_sheet.cell(1, 1)
    sunshine_title.value = "【七八联动 燃聚‘滇’峰】竞赛追踪"
    sunshine_title.fill = PatternFill("solid", fgColor="203864")
    sunshine_title.font = Font(color="FFFFFF", bold=True, size=13)
    sunshine_title.alignment = Alignment(horizontal="center", vertical="center")
    sunshine_sheet.row_dimensions[1].height = 24
    sunshine_sheet.freeze_panes = "A3"
    sunshine_sheet.auto_filter.ref = f"A2:I{sunshine_sheet.max_row}"

    for row in sunshine_sheet.iter_rows(min_row=2, max_row=sunshine_sheet.max_row, max_col=9):
        is_header = row[0].row == 2
        is_qualified = row[8].value == "达标"
        for cell in row:
            cell.border = table_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if is_header:
                cell.fill = PatternFill("solid", fgColor="203864")
                cell.font = Font(color="FFFFFF", bold=True)
            elif is_qualified:
                cell.fill = blue_fill
            else:
                cell.fill = white_fill
            if cell.column in {7, 8} and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"
            if cell.column == 9:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0"
                    if cell.value < 0:
                        cell.font = Font(color="FF0000", bold=True)
                elif cell.value == "达标":
                    cell.font = Font(bold=True)

    for column_index, width in enumerate([8, 14, 28, 20, 14, 14, 10, 14, 14], start=1):
        sunshine_sheet.column_dimensions[get_column_letter(column_index)].width = width

    sunshine_summary = reports.get("sunshine_region_summary")
    if isinstance(sunshine_summary, pd.DataFrame):
        start_col = 11
        start_row = 2
        headers = ["序号", "区域", "符合参赛人数", "达标人数", "未达标人数"]
        for offset, header in enumerate(headers):
            cell = sunshine_sheet.cell(start_row, start_col + offset)
            cell.value = header
            cell.fill = PatternFill("solid", fgColor="002060" if header != "未达标人数" else "C00000")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = table_border

        for index, row_data in sunshine_summary.iterrows():
            excel_row = start_row + index + 1
            values = [
                row_data["序号"],
                row_data["区域"],
                int(row_data["符合参赛人数"]),
                int(row_data["达标人数"]),
                int(row_data["未达标人数"]),
            ]
            is_total = row_data["序号"] == "总计"
            for offset, value in enumerate(values):
                cell = sunshine_sheet.cell(excel_row, start_col + offset)
                cell.value = value
                cell.border = table_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if is_total:
                    cell.fill = PatternFill("solid", fgColor="002060" if offset < 4 else "C00000")
                    cell.font = Font(color="FFFFFF", bold=True)
                elif offset == 4:
                    cell.font = Font(color="FF0000")

        for offset, width in enumerate([8, 24, 16, 14, 14]):
            sunshine_sheet.column_dimensions[get_column_letter(start_col + offset)].width = width

    sunshine_plan_image = ExcelImage(str(write_embedded_image(SUNSHINE_PLAN_IMAGE_BASE64, "sunshine_plan.png")))
    sunshine_plan_image.width = 610
    sunshine_plan_image.height = 352
    sunshine_sheet.add_image(sunshine_plan_image, "K23")

    if "（信泰）7月专属方案" in workbook.sheetnames:
        xintai_sheet = workbook["（信泰）7月专属方案"]
        xintai_sheet.insert_rows(1)
        xintai_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
        xintai_title = xintai_sheet.cell(1, 1)
        xintai_title.value = "夏日狂欢·培训悦游 信泰7月专属方案"
        xintai_title.fill = PatternFill("solid", fgColor="203864")
        xintai_title.font = Font(color="FFFFFF", bold=True, size=13)
        xintai_title.alignment = Alignment(horizontal="center", vertical="center")
        xintai_sheet.row_dimensions[1].height = 24
        xintai_sheet.freeze_panes = "A3"
        xintai_sheet.auto_filter.ref = f"A2:I{xintai_sheet.max_row}"
    
        for row in xintai_sheet.iter_rows(min_row=2, max_row=xintai_sheet.max_row, max_col=9):
            is_header = row[0].row == 2
            is_qualified = row[8].value == "达标"
            for cell in row:
                cell.border = table_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if is_header:
                    cell.fill = PatternFill("solid", fgColor="203864")
                    cell.font = Font(color="FFFFFF", bold=True)
                elif is_qualified:
                    cell.fill = blue_fill
                else:
                    cell.fill = white_fill
                if cell.column in {7, 8} and isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0"
                if cell.column == 9:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = "#,##0"
                        if cell.value < 0:
                            cell.font = Font(color="FF0000", bold=True)
                    elif cell.value == "达标":
                        cell.font = Font(color="FF0000", bold=True)
    
        for column_index, width in enumerate([8, 14, 28, 20, 14, 14, 10, 14, 14], start=1):
            xintai_sheet.column_dimensions[get_column_letter(column_index)].width = width
    
        xintai_summary = reports.get("xintai_region_summary")
        if isinstance(xintai_summary, pd.DataFrame):
            start_col = 13
            start_row = 2
            headers = ["序号", "区域", "符合参赛人数", "达标人数", "未达标人数"]
            for offset, header in enumerate(headers):
                cell = xintai_sheet.cell(start_row, start_col + offset)
                cell.value = header
                cell.fill = PatternFill("solid", fgColor="002060" if header != "未达标人数" else "C00000")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = table_border
    
            for index, row_data in xintai_summary.iterrows():
                excel_row = start_row + index + 1
                values = [
                    row_data["序号"],
                    row_data["区域"],
                    int(row_data["符合参赛人数"]),
                    int(row_data["达标人数"]),
                    int(row_data["未达标人数"]),
                ]
                is_total = row_data["序号"] == "总计"
                for offset, value in enumerate(values):
                    cell = xintai_sheet.cell(excel_row, start_col + offset)
                    cell.value = value
                    cell.border = table_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if is_total:
                        cell.fill = PatternFill("solid", fgColor="002060" if offset < 4 else "C00000")
                        cell.font = Font(color="FFFFFF", bold=True)
                    elif offset == 4:
                        cell.font = Font(color="FF0000")
    
            for offset, width in enumerate([8, 24, 16, 14, 14]):
                xintai_sheet.column_dimensions[get_column_letter(start_col + offset)].width = width
    
        xintai_plan_image = ExcelImage(str(write_embedded_image(XINTAI_PLAN_IMAGE_BASE64, "xintai_plan.png")))
        xintai_plan_image.width = 610
        xintai_plan_image.height = 351
        xintai_sheet.add_image(xintai_plan_image, "M23")
    
    mdrt_sheet = workbook["MDRT"]
    mdrt_sheet.insert_rows(1)
    mdrt_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    mdrt_title = mdrt_sheet.cell(1, 1)
    mdrt_title.value = f"全年【百万精英 浪漫之旅】--个人通道竞赛差距"
    mdrt_title.fill = PatternFill("solid", fgColor="203864")
    mdrt_title.font = Font(color="FFFFFF", bold=True, size=14)
    mdrt_title.alignment = Alignment(horizontal="center", vertical="center")
    mdrt_sheet.row_dimensions[1].height = 24
    mdrt_sheet.freeze_panes = "A3"
    mdrt_sheet.auto_filter.ref = f"A2:H{mdrt_sheet.max_row}"

    for row in mdrt_sheet.iter_rows(min_row=2, max_row=mdrt_sheet.max_row, max_col=8):
        is_header = row[0].row == 2
        is_top = row[0].row in {3, 4}
        for cell in row:
            cell.border = table_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if is_header:
                cell.fill = PatternFill("solid", fgColor="203864")
                cell.font = Font(color="FFFFFF", bold=True)
            elif is_top:
                cell.fill = blue_fill
                if cell.column in {2, 6}:
                    cell.font = Font(bold=True)
            else:
                cell.fill = white_fill
                if cell.column in {2, 6}:
                    cell.font = Font(bold=True)
            if cell.column == 8 and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"

    for column_index, width in enumerate([8, 18, 14, 28, 22, 14, 14, 18], start=1):
        mdrt_sheet.column_dimensions[get_column_letter(column_index)].width = width

    mdrt_region_sheet = workbook["MDRT-区域"]
    mdrt_region_sheet.insert_rows(1)
    mdrt_region_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    mdrt_region_title = mdrt_region_sheet.cell(1, 1)
    mdrt_region_title.value = "总监区域业务进度"
    mdrt_region_title.fill = PatternFill("solid", fgColor="203864")
    mdrt_region_title.font = Font(color="FFFFFF", bold=True, size=16)
    mdrt_region_title.alignment = Alignment(horizontal="center", vertical="center")
    mdrt_region_sheet.row_dimensions[1].height = 30
    mdrt_region_sheet.freeze_panes = "A3"
    mdrt_region_sheet.auto_filter.ref = f"A2:C{mdrt_region_sheet.max_row}"

    for row in mdrt_region_sheet.iter_rows(min_row=2, max_row=mdrt_region_sheet.max_row, max_col=3):
        is_header = row[0].row == 2
        for cell in row:
            cell.border = table_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if is_header:
                cell.fill = PatternFill("solid", fgColor="203864")
                cell.font = Font(color="FFFFFF", bold=True, size=14)
            else:
                cell.fill = white_fill
            if cell.column == 3 and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"

    for column_index, width in enumerate([10, 28, 18], start=1):
        mdrt_region_sheet.column_dimensions[get_column_letter(column_index)].width = width

    local_agent_sheet = workbook["大湾区属地代理人达成情况"]
    local_agent_sheet.insert_rows(1)
    local_agent_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    local_title = local_agent_sheet.cell(1, 1)
    local_title.value = "大湾区属地代理人达成情况"
    local_title.fill = PatternFill("solid", fgColor="B4C6E7")
    local_title.font = Font(color="000000", bold=True, size=16)
    local_title.alignment = Alignment(horizontal="center", vertical="center")
    local_agent_sheet.row_dimensions[1].height = 24
    local_agent_sheet.freeze_panes = "A3"
    local_agent_sheet.auto_filter.ref = f"A2:I{local_agent_sheet.max_row}"

    for row in local_agent_sheet.iter_rows(min_row=2, max_row=local_agent_sheet.max_row, max_col=9):
        is_header = row[0].row == 2
        for cell in row:
            cell.border = table_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if is_header:
                cell.fill = PatternFill("solid", fgColor="4472C4")
                cell.font = Font(color="FFFFFF", bold=True)
            else:
                cell.fill = PatternFill("solid", fgColor="D9E6F2")
            if cell.column in {6, 7} and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"
                if cell.column == 7 and cell.value < 0:
                    cell.font = Font(color="000000")

    for column_index, width in enumerate([8, 22, 20, 16, 14, 14, 16, 16, 14], start=1):
        local_agent_sheet.column_dimensions[get_column_letter(column_index)].width = width

    local_summary = reports.get("local_agent_region_summary")
    if isinstance(local_summary, pd.DataFrame):
        start_col = 11
        start_row = 2
        headers = ["序号", "区域", "考核人数", "已达标人数", "未达标人数"]
        for offset, header in enumerate(headers):
            cell = local_agent_sheet.cell(start_row, start_col + offset)
            cell.value = header
            cell.fill = PatternFill("solid", fgColor="4472C4" if header != "未达标人数" else "C65911")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = table_border

        for index, row_data in local_summary.iterrows():
            excel_row = start_row + index + 1
            values = [
                row_data["序号"],
                row_data["区域"],
                int(row_data["考核人数"]),
                int(row_data["已达标人数"]),
                int(row_data["未达标人数"]),
            ]
            is_total = row_data["序号"] == "合计"
            for offset, value in enumerate(values):
                cell = local_agent_sheet.cell(excel_row, start_col + offset)
                cell.value = value
                cell.border = table_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if is_total:
                    cell.fill = PatternFill("solid", fgColor="4472C4" if offset < 4 else "C65911")
                    cell.font = Font(color="FFFFFF", bold=True)
                elif offset == 4:
                    cell.font = Font(color="000000")

        for offset, width in enumerate([8, 24, 14, 14, 14]):
            local_agent_sheet.column_dimensions[get_column_letter(start_col + offset)].width = width

    workbook.save(path)

def export_excel(
    reports: dict[str, pd.DataFrame | dict[str, float] | pd.Series],
    output_path: Path,
    dashboard_png: Path,
    report_date: pd.Timestamp,
) -> None:
    rules = pd.DataFrame(
        [
            ["报告日期", report_date.strftime("%Y-%m-%d")],
            ["数据文件", str(INPUT_FILE) if INPUT_FILE else f"自动读取 {DATA_FOLDER}/{FILE_PATTERN} 中最新文件"],
            ["团队筛选", f"出单团队包含‘{TEAM_KEYWORD}’"],
            ["绩效排除", "最开始剔除撤销、犹豫期退保、未承保/待承保、空白状态；绩效中继续排除撤单"],
            ["达标门槛", PVI_THRESHOLD],
            ["临近门槛", f"{NEAR_THRESHOLD_MIN:,.0f} ≤ 本月PVI < {PVI_THRESHOLD:,.0f}"],
            ["尚未出单口径", f"最近{ROSTER_LOOKBACK_DAYS}天出现过的代理，且本月没有正PVI"],
            ["连续未出单", f"距离最近正PVI承保日期不少于{INACTIVE_DAYS}天"],
            ["超3天未承保", "交单日期已超过3天且承保日期为空"],
            ["正常业务预测", "剔除本月最大一笔PVI后，按已过工作日日均推算整月"],
            ["月度目标", MONTHLY_TARGET],
            ["目标缺口", "月度目标减本月净PVI；若已达标则为0"],
            ["预测工作日", "按周一至周五计算，未自动识别中国法定节假日调休"],
            ["注意", "精确识别本月尚未出单代理，建议提供正式在职代理花名册"],
        ],
        columns=["规则", "定义"],
    )
    frames = {
        "7月安盛活力星达成追踪": reports["competition_tracking"],
        "(阳光)七八联动七月追踪": reports["sunshine_tracking"],
        "（信泰）7月专属方案": reports["xintai_tracking"],
        "MDRT": reports["mdrt_tracking"],
        "MDRT-区域": reports["mdrt_region_summary"],
        "大湾区属地代理人达成情况": reports["local_agent_tracking"],
    }
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, frame in frames.items():
            assert isinstance(frame, pd.DataFrame)
            frame.to_excel(writer, sheet_name=sheet_name, index=False)
    style_excel(output_path, dashboard_png, reports)
