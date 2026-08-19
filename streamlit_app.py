from __future__ import annotations

import tempfile
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ============================================================
# 每天只需要检查这里
# ============================================================
from pathlib import Path

# 固定分析年份
ANALYSIS_YEAR = 2026

# 月份配置：以后切换月份时优先改这里
REPORT_MONTH = 8
REPORT_MONTH_LABEL = f"{REPORT_MONTH}月"
CUMULATIVE_MONTHS = [7, 8]
CUMULATIVE_MONTH_LABEL = "-".join(str(month) for month in CUMULATIVE_MONTHS) + "月"


# 下面这些口径一般不需要每天修改。
SHEET_NAME = "保单汇总列表"
TEAM_KEYWORD = "大湾区计划"

PVI_THRESHOLD = 6000.0
SUNSHINE_COMPANY_KEYWORD = "阳光人寿"
SUNSHINE_THRESHOLD = 100_000.0
MDRT_REGION_TARGET = 5_000_000.0
LOCAL_AGENT_THRESHOLD = 20_000.0

SHEET_NAMES = {
    "tracking": f"{REPORT_MONTH_LABEL}安盛活力星达成追踪",
    "sunshine": "(阳光)七八联动追踪",
    "mdrt": "MDRT",
    "mdrt_region": "MDRT-区域",
    "local_agent": "大湾区属地代理人达成情况",
    "region_month_pvi": f"区域{CUMULATIVE_MONTH_LABEL}PVI汇总",
    "agency_team_pvi": f"机构团队{REPORT_MONTH_LABEL}PVI汇总",
    "agency_team_hires": f"机构团队{REPORT_MONTH_LABEL}入职人数",
    "kai_xuan_detail": "大湾区凯旋明细（全年）",
    "yaokun_detail": "曜坤区明细（全年）",
    "fangyuan_hq_detail": "方圆区总部直辖明细（全年）",
}

TRACKING_SHEET_NAME = SHEET_NAMES["tracking"]
SUNSHINE_SHEET_NAME = SHEET_NAMES["sunshine"]
REGION_MONTH_PVI_SHEET_NAME = SHEET_NAMES["region_month_pvi"]
AGENCY_TEAM_PVI_SHEET_NAME = SHEET_NAMES["agency_team_pvi"]
AGENCY_TEAM_HIRES_SHEET_NAME = SHEET_NAMES["agency_team_hires"]

FINAL_SHEETS = list(SHEET_NAMES.values())

CUMULATIVE_PVI_COLUMN = f"{CUMULATIVE_MONTH_LABEL}合计PVI"
REPORT_MONTH_PVI_COLUMN = f"{REPORT_MONTH_LABEL}总PVI"
REPORT_MONTH_HIRES_COLUMN = f"{REPORT_MONTH_LABEL}入职人数"

MONTH_PVI_COLUMNS = [f"{month}月总PVI" for month in CUMULATIVE_MONTHS]
MONTH_PVI_MAP = dict(zip(CUMULATIVE_MONTHS, MONTH_PVI_COLUMNS))

REQUIRED_COLUMNS = [
    "出单团队",
    "出单机构",
    "出单人所属区域",
    "出单人所属组织",
    "管理所属区域",
    "管理所属组",
    "服务代理人姓名",
    "服务代理人工号",
    "出单代理人工号",
    "出单代理人姓名",
    "保单号",
    "投保单号",
    "PVI",
    "交单日期",
    "承保日期",
    "承保进度",
    "保单状态",
    "险种名称",
    "产品统计分类",
    "保险公司",
    "回执日期",
    "回访完成日期",
]

EXCLUDED_POLICY_STATUSES = {
    "撤销",
    "犹豫期退保",
    "未承保",
    "待承保",
}

MDRT_LONG_TERM_LIFE_CATEGORIES = {
    "定期寿险",
    "两全险",
    "年金险",
    "终身寿险",
    "重疾险",
}

LOCAL_AGENT_ROSTER = [
    {"所属区域": "丰盛区", "所属组织": "伍尚明直辖组", "工号": "AXA004090", "代理人姓名": "王诗婕", "考核截止时间": "2026年8月31日", "入职日期": "2026-05-18"},
    {"所属区域": "SCREM 关爱区", "所属组织": "帅玉芬直辖组", "工号": "AXA004042", "代理人姓名": "余素娴", "考核截止时间": "2026年9月30日", "入职日期": "2026-06-23"},
    {"所属区域": "SCREM 关爱区", "所属组织": "卢海玉直辖组", "工号": "AXA004121", "代理人姓名": "黄丽枝", "考核截止时间": "2026年9月30日", "入职日期": "2026-06-11"},
    {"所属区域": "SCREM 关爱区", "所属组织": "卢海玉直辖组", "工号": "AXA004141", "代理人姓名": "叶伟灿", "考核截止时间": "2026年9月30日", "入职日期": "2026-06-29"},
    {"所属区域": "蓝天区", "所属组织": "朱申萍直辖组", "工号": "AXA004106", "代理人姓名": "段远婷", "考核截止时间": "2026年9月30日", "入职日期": "2026-06-30"},
    {"所属区域": "SCREM 关爱区", "所属组织": "卢海玉直辖组", "工号": "AXA004139", "代理人姓名": "叶小梨", "考核截止时间": "2026年10月31日", "入职日期": "2026-07-02"},
    {"所属区域": "SCREM 关爱区", "所属组织": "卢海玉直辖组", "工号": "AXA004144", "代理人姓名": "黄伟烈", "考核截止时间": "2026年10月31日", "入职日期": "2026-07-02"},
    {"所属区域": "SCREM 关爱区", "所属组织": "卢海玉直辖组", "工号": "AXA004142", "代理人姓名": "叶鹏娜", "考核截止时间": "2026年10月31日", "入职日期": "2026-07-02"},
    {"所属区域": "SCREM 关爱区", "所属组织": "卢海玉直辖组", "工号": "AXA004165", "代理人姓名": "卢惠玲", "考核截止时间": "2026年10月31日", "入职日期": "2026-07-08"},
    {"所属区域": "SCREM 关爱区", "所属组织": "卢海玉直辖组", "工号": "AXA004140", "代理人姓名": "黄后仪", "考核截止时间": "2026年10月31日", "入职日期": "2026-07-15"},
    {"所属区域": "SCREM 关爱区", "所属组织": "卢海玉直辖组", "工号": "AXA004178", "代理人姓名": "陈浩棠", "考核截止时间": "2026年10月31日", "入职日期": "2026-07-15"},
    {"所属区域": "旭日&旭日北辰区", "所属组织": "林辰龙直辖组", "工号": "AXA003839", "代理人姓名": "黄旭娜", "考核截止时间": "2026年10月31日", "入职日期": "2026-07-16"},
    {"所属区域": "SCREM 关爱区", "所属组织": "卢海玉直辖组", "工号": "AXA004205", "代理人姓名": "郑翠雯", "考核截止时间": "2026年10月31日", "入职日期": "2026-07-24"},
    {"所属区域": "SCREM 关爱区", "所属组织": "卢海玉直辖组", "工号": "AXA004158", "代理人姓名": "余思敏", "考核截止时间": "2026年10月31日", "入职日期": "2026-07-24"},
    {"所属区域": "SCREM 关爱区", "所属组织": "卢海玉直辖组", "工号": "AXA004138", "代理人姓名": "孙敬珉", "考核截止时间": "2026年10月31日", "入职日期": "2026-07-24"},
    {"所属区域": "SCREM 关爱区", "所属组织": "卢海玉直辖组", "工号": "AXA004211", "代理人姓名": "陈咏珊", "考核截止时间": "2026年10月31日", "入职日期": "2026-07-27"},
    {"所属区域": "SCREM 关爱区", "所属组织": "卢海玉直辖组", "工号": "AXA004206", "代理人姓名": "王燕仪", "考核截止时间": "2026年10月31日", "入职日期": "2026-07-27"},
    {"所属区域": "SCREM 关爱区", "所属组织": "卢海玉直辖组", "工号": "AXA004184", "代理人姓名": "刘燕婷", "考核截止时间": "2026年10月31日", "入职日期": "2026-07-27"},
    {"所属区域": "旭日&旭日北辰区", "所属组织": "林辰龙直辖组", "工号": "AXA004218", "代理人姓名": "李育萍", "考核截止时间": "2026年11月30日", "入职日期": "2026-08-03"},
    {"所属区域": "旭日&旭日北辰区", "所属组织": "林辰龙直辖组", "工号": "AXA004219", "代理人姓名": "张媄淇", "考核截止时间": "2026年11月30日", "入职日期": "2026-08-04"},
]

def latest_source_file() -> Path:
    selected = SELECTED_INPUT_FILE if "SELECTED_INPUT_FILE" in globals() else find_source_file()
    print(f"本次分析的表格：{selected.name}")
    print(f"表格完整路径：{selected}")
    return selected


def text_series(series: pd.Series) -> pd.Series:
    return series.astype("string").str.strip()


def parse_pvi(series: pd.Series) -> pd.Series:
    cleaned = text_series(series).str.replace(r"[,，￥¥\s]", "", regex=True)
    return pd.to_numeric(cleaned, errors="coerce").fillna(0.0)



def money(value: float) -> str:
    return f"{value:,.2f}".rstrip("0").rstrip(".")

def empty_frame(columns: list[str]) -> pd.DataFrame:
    return pd.DataFrame(columns=columns)


def normalize_region_name(value: object) -> str:
    text = "" if pd.isna(value) else str(value).strip()
    if not text:
        return "未填写"

    text = text.split("-", 1)[0].strip()

    aliases = {
        "Bees 旭日区": "旭日&旭日北辰区",
        "Bees 旭日北辰区": "旭日&旭日北辰区",
    }

    return aliases.get(text, text)



def clean_data(path: Path, sheet: str, team_keyword: str) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name=sheet, dtype=object)
    df.columns = df.columns.astype(str).str.strip()

    missing = [column for column in REQUIRED_COLUMNS if column not in df.columns]
    if missing:
        raise ValueError(f"工作表缺少必要字段：{missing}")

    df = df.copy()
    df["_pvi"] = parse_pvi(df["PVI"])

    # 全局口径修正：陕西分公司统一归入大湾区计划。
    df.loc[
        text_series(df["出单机构"]).eq("陕西分公司"),
        "出单团队",
    ] = team_keyword

    # 全局口径修正：虚拟总监陕西使用管理/服务字段替换出单字段。
    virtual_shaanxi_mask = (
        text_series(df["出单人所属区域"]).eq("内部测试区域")
        & text_series(df["出单代理人姓名"]).eq("虚拟总监陕西")
    )

    replace_columns = {
        "出单人所属区域": "管理所属区域",
        "出单人所属组织": "管理所属组",
        "出单代理人姓名": "服务代理人姓名",
        "出单代理人工号": "服务代理人工号",
    }

    for target, source in replace_columns.items():
        df.loc[virtual_shaanxi_mask, target] = df.loc[virtual_shaanxi_mask, source]

    date_columns = {
        "交单日期": "_submit_date",
        "承保日期": "_insured_date",
        "回执日期": "_receipt_date",
        "回访完成日期": "_visit_date",
    }

    for source, target in date_columns.items():
        df[target] = pd.to_datetime(df[source], errors="coerce").dt.normalize()

    text_columns = {
        "出单团队": "_team",
        "保单状态": "_status",
        "承保进度": "_progress",
        "出单机构": "_agency",
        "出单人所属区域": "_region",
        "出单人所属组织": "_organization",
        "管理所属区域": "_management_region",
        "管理所属组": "_management_organization",
        "出单代理人工号": "_agent_id",
        "出单代理人姓名": "_agent_name",
        "产品统计分类": "_product_category",
    }

    for source, target in text_columns.items():
        df[target] = text_series(df[source])

    valid_status = (
        df["_status"].notna()
        & df["_status"].ne("")
        & ~df["_status"].isin(EXCLUDED_POLICY_STATUSES)
        & ~df["_status"].str.contains("撤销|犹豫期退保", na=False)
    )

    df = df.loc[valid_status].copy()
    df["_is_team"] = df["_team"].eq(team_keyword)

    # 不在清洗阶段提前筛掉其他团队，后续不同 sheet 会各自筛选。
    df["_is_cancelled"] = df["_status"].str.contains("犹豫期退保|撤销", na=False)
    df["_is_valid"] = ~df["_is_cancelled"]

    return df


def determine_report_date(
    df: pd.DataFrame,
    requested: str | None
) -> pd.Timestamp:

    if requested:
        parsed = pd.to_datetime(requested, errors="raise")
        parsed = pd.Timestamp(parsed).normalize()

        if parsed.year != ANALYSIS_YEAR:
            raise ValueError(
                f"报告日期必须属于{ANALYSIS_YEAR}年，"
                f"当前填写的是：{parsed:%Y-%m-%d}"
            )

        return parsed

    latest = df.loc[
        df["_is_team"]
        & df["_insured_date"].dt.year.eq(ANALYSIS_YEAR),
        "_insured_date"
    ].max()

    if pd.isna(latest):
        raise ValueError(
            f"{ANALYSIS_YEAR}年大湾区数据中没有可用的承保日期"
        )

    return pd.Timestamp(latest).normalize()

def style_excel(
    path: Path,
    reports: dict[str, pd.DataFrame | dict[str, int]],
) -> None:
    from openpyxl import load_workbook

    workbook = load_workbook(path)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin_gray = Side(style="thin", color="D9E1F2")

    blue_fill = PatternFill("solid", fgColor="B4C6E7")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    black_side = Side(style="thin", color="000000")
    table_border = Border(
        left=black_side,
        right=black_side,
        top=black_side,
        bottom=black_side,
    )

    def add_title(
        sheet,
        title: str,
        end_column: int,
        fill_color: str = "203864",
        font_color: str = "FFFFFF",
        font_size: int = 14,
        height: int = 24,
    ) -> None:
        sheet.insert_rows(1)
        sheet.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=end_column,
        )

        cell = sheet.cell(1, 1)
        cell.value = title
        cell.fill = PatternFill("solid", fgColor=fill_color)
        cell.font = Font(color=font_color, bold=True, size=font_size)
        cell.alignment = Alignment(horizontal="center", vertical="center")

        sheet.row_dimensions[1].height = height
        sheet.freeze_panes = "A3"
        sheet.auto_filter.ref = f"A2:{get_column_letter(end_column)}{sheet.max_row}"

    def style_table(
        sheet,
        max_col: int,
        header_color: str = "203864",
        highlight_col: int | None = None,
        highlight_value: str = "达标",
        body_fill: str = "FFFFFF",
    ) -> None:
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=max_col):
            is_header = row[0].row == 2
            is_highlight = (
                highlight_col is not None
                and row[highlight_col - 1].value == highlight_value
            )

            for cell in row:
                cell.border = table_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

                if is_header:
                    cell.fill = PatternFill("solid", fgColor=header_color)
                    cell.font = Font(color="FFFFFF", bold=True)
                elif is_highlight:
                    cell.fill = blue_fill
                else:
                    cell.fill = PatternFill("solid", fgColor=body_fill)

    def set_widths(sheet, widths: list[int], start: int = 1) -> None:
        for offset, width in enumerate(widths):
            sheet.column_dimensions[get_column_letter(start + offset)].width = width

    def write_summary_table(
        sheet,
        summary: pd.DataFrame,
        start_col: int,
        start_row: int,
        headers: list[str],
        total_label: str = "总计",
        red_last_col: bool = False,
    ) -> None:
        for offset, header in enumerate(headers):
            cell = sheet.cell(start_row, start_col + offset)
            cell.value = header
            cell.fill = PatternFill(
                "solid",
                fgColor="C00000" if red_last_col and offset == len(headers) - 1 else "4472C4",
            )
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = table_border

        for index, row_data in summary.iterrows():
            excel_row = start_row + index + 1
            is_total = str(row_data[headers[0]]) == total_label

            for offset, header in enumerate(headers):
                cell = sheet.cell(excel_row, start_col + offset)
                cell.value = row_data[header]
                cell.border = table_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

                if is_total:
                    cell.fill = PatternFill(
                        "solid",
                        fgColor="C00000" if red_last_col and offset == len(headers) - 1 else "4472C4",
                    )
                    cell.font = Font(color="FFFFFF", bold=True)
                elif red_last_col and offset == len(headers) - 1:
                    cell.font = Font(color="FF0000")

    # 只保留最终需要输出的 sheet。
    for sheet in list(workbook.worksheets):
        if sheet.title not in FINAL_SHEETS:
            workbook.remove(sheet)

    # 通用基础格式。
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.sheet_view.showGridLines = False

        if sheet.max_row >= 1 and sheet.max_column >= 1:
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(bottom=thin_gray)

            sheet.auto_filter.ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"

        for column_index in range(1, sheet.max_column + 1):
            values = [
                sheet.cell(row, column_index).value
                for row in range(1, min(sheet.max_row, 200) + 1)
            ]
            max_length = max(
                (len(str(value)) for value in values if value is not None),
                default=8,
            )
            sheet.column_dimensions[get_column_letter(column_index)].width = min(
                max(max_length + 2, 10),
                35,
            )

        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if cell.is_date:
                    cell.number_format = "yyyy-mm-dd"
                elif isinstance(cell.value, float):
                    cell.number_format = "#,##0.00"

    # 安盛活力星达成追踪
    tracking = workbook[TRACKING_SHEET_NAME]
    add_title(tracking, TRACKING_SHEET_NAME, end_column=8)
    style_table(tracking, max_col=8, highlight_col=8)
    set_widths(tracking, [8, 14, 28, 20, 14, 14, 14, 12])

    for row in tracking.iter_rows(min_row=2, max_row=tracking.max_row, max_col=8):
        for cell in row:
            if cell.column == 7 and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"
            if cell.column == 8:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0"
                    cell.font = Font(bold=True)
                elif cell.value == "达标":
                    cell.font = Font(bold=True)

    region_summary = reports.get("qualified_region_summary")
    if isinstance(region_summary, pd.DataFrame):
        write_summary_table(
            tracking,
            region_summary,
            start_col=11,
            start_row=2,
            headers=["区域", "达标数"],
            total_label="总计",
        )
        set_widths(tracking, [28, 12], start=11)

    # 阳光联动
    sunshine_sheet = workbook[SUNSHINE_SHEET_NAME]
    add_title(
        sunshine_sheet,
        "【七八联动 燃聚‘滇’峰】竞赛追踪",
        end_column=9,
        font_size=13,
    )
    style_table(sunshine_sheet, max_col=9, highlight_col=9)
    set_widths(sunshine_sheet, [8, 14, 28, 20, 14, 14, 10, 14, 14])

    for row in sunshine_sheet.iter_rows(min_row=2, max_row=sunshine_sheet.max_row, max_col=9):
        for cell in row:
            if cell.column in {7, 8} and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"
            if cell.column == 9:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0"
                    if cell.value < 0:
                        cell.font = Font(color="FF0000", bold=True)
                elif cell.value == "达标":
                    cell.font = Font(bold=True)

    sunshine_summary = reports.get("sunshine_region_summary")
    if isinstance(sunshine_summary, pd.DataFrame):
        write_summary_table(
            sunshine_sheet,
            sunshine_summary,
            start_col=11,
            start_row=2,
            headers=["序号", "区域", "符合参赛人数", "达标人数", "未达标人数"],
            total_label="总计",
            red_last_col=True,
        )
        set_widths(sunshine_sheet, [8, 24, 16, 14, 14], start=11)

    # MDRT
    mdrt_sheet = workbook["MDRT"]
    add_title(
        mdrt_sheet,
        "全年【百万精英 浪漫之旅】--个人通道竞赛差距",
        end_column=10,
    )
    style_table(mdrt_sheet, max_col=10)
    set_widths(mdrt_sheet, [8, 18, 14, 28, 22, 14, 14, 18, 18, 12])

    for row in mdrt_sheet.iter_rows(min_row=2, max_row=mdrt_sheet.max_row, max_col=10):
        is_header = row[0].row == 2
        for cell in row:
            if not is_header and cell.column in {2, 6}:
                cell.font = Font(bold=True)
            if cell.column in {8, 9} and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"
            if cell.column == 10 and isinstance(cell.value, (int, float)):
                cell.number_format = "0.00%"

    # MDRT-区域
    mdrt_region_sheet = workbook["MDRT-区域"]
    add_title(
        mdrt_region_sheet,
        "总监区域业务进度",
        end_column=4,
        font_size=16,
        height=30,
    )
    style_table(mdrt_region_sheet, max_col=4)
    set_widths(mdrt_region_sheet, [10, 18, 28, 18])

    for row in mdrt_region_sheet.iter_rows(min_row=2, max_row=mdrt_region_sheet.max_row, max_col=4):
        for cell in row:
            if cell.column == 4 and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"

    # 大湾区属地代理人达成情况
    local_agent_sheet = workbook["大湾区属地代理人达成情况"]
    add_title(
        local_agent_sheet,
        "大湾区属地代理人达成情况",
        end_column=9,
        fill_color="B4C6E7",
        font_color="000000",
        font_size=16,
    )
    style_table(
        local_agent_sheet,
        max_col=9,
        header_color="4472C4",
        body_fill="D9E6F2",
    )
    set_widths(local_agent_sheet, [8, 22, 20, 16, 14, 14, 16, 16, 14])

    for row in local_agent_sheet.iter_rows(min_row=2, max_row=local_agent_sheet.max_row, max_col=9):
        for cell in row:
            if cell.column in {6, 7} and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"
                if cell.column == 7 and cell.value < 0:
                    cell.font = Font(color="000000")

    local_summary = reports.get("local_agent_region_summary")
    if isinstance(local_summary, pd.DataFrame):
        write_summary_table(
            local_agent_sheet,
            local_summary,
            start_col=11,
            start_row=2,
            headers=["序号", "区域", "考核人数", "已达标人数", "未达标人数"],
            total_label="合计",
            red_last_col=True,
        )
        set_widths(local_agent_sheet, [8, 24, 14, 14, 14], start=11)

    workbook.save(path)


def export_excel(
    reports: dict[str, pd.DataFrame | dict[str, int]],
    output_path: Path,
    report_date: pd.Timestamp,
) -> None:
    rules = pd.DataFrame(
        [
            ["报告日期", report_date.strftime("%Y-%m-%d")],
            ["数据文件", str(INPUT_FILE) if INPUT_FILE else f"自动读取 {DATA_FOLDER}/{FILE_PATTERN} 中最新文件"],
            ["有效保单状态", "剔除撤销、犹豫期退保、未承保/待承保、空白状态"],
            ["安盛活力星", f"出单团队为{TEAM_KEYWORD}，承保月份为{REPORT_MONTH_LABEL}，按工号汇总PVI并判断是否达到{PVI_THRESHOLD:,.0f}"],
            ["阳光联动", f"出单团队为{TEAM_KEYWORD}，保险公司包含阳光人寿，承保月份属于{CUMULATIVE_MONTH_LABEL}，合计PVI达到{SUNSHINE_THRESHOLD:,.0f}为达标"],
            ["MDRT", f"统计{ANALYSIS_YEAR}年所有有效承保PVI，不限制出单团队；MDRT-区域基于MDRT结果汇总"],
            ["区域累计PVI", f"按所属团队和出单人所属区域汇总{CUMULATIVE_MONTH_LABEL}PVI，并按当前维护规则合并部分区域名称"],
            ["代理人信息表", "用于机构团队月度入职人数、今日新增、本月累计新增，以及MDRT人员归属修正"],
        ],
        columns=["规则", "定义"],
    )

    frames = {
        TRACKING_SHEET_NAME: reports["competition_tracking"],
        SUNSHINE_SHEET_NAME: reports["sunshine_tracking"],
        "MDRT": reports["mdrt_tracking"],
        "MDRT-区域": reports["mdrt_region_summary"],
        "大湾区属地代理人达成情况": reports["local_agent_tracking"],
        REGION_MONTH_PVI_SHEET_NAME: reports["region_july_august_pvi"],
        AGENCY_TEAM_PVI_SHEET_NAME: reports["agency_team_august_pvi"],
        AGENCY_TEAM_HIRES_SHEET_NAME: reports["agency_team_august_hires"],
        "大湾区凯旋明细（全年）": reports["kai_xuan_detail"],
        "曜坤区明细（全年）": reports["yaokun_detail"],
        "方圆区总部直辖明细（全年）": reports["fangyuan_hq_detail"],
    }

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, frame in frames.items():
            assert isinstance(frame, pd.DataFrame)
            frame.to_excel(writer, sheet_name=sheet_name, index=False)

    style_excel(output_path, reports)

st.set_page_config(page_title="PVI自动化报表Demo", page_icon="📊", layout="wide")
st.title("保险经销业务 PVI 自动化报表分析工具")
st.caption("上传保单汇总列表和代理人信息表，自动生成新版多 Sheet Excel 报告。")

with st.sidebar:
    st.header("统计配置")
    report_month = st.number_input("报告月份", min_value=1, max_value=12, value=REPORT_MONTH, step=1)
    cumulative_months = st.multiselect(
        "累计统计月份",
        options=list(range(1, 13)),
        default=CUMULATIVE_MONTHS,
    )
    report_date_text = st.text_input("补算日期（可选，格式 YYYY-MM-DD）", value="")
    st.caption("留空时，系统自动使用数据中 2026 年大湾区计划最新承保日期。")

policy_file = st.file_uploader("上传保单汇总列表 Excel", type=["xlsx"])
agent_file = st.file_uploader("上传代理人信息 Excel", type=["xlsx", "xls"])

st.info(
    "当前版本已移除图片看板、月度目标缺口、目标达成率和信泰旧方案；"
    "保留安盛活力星、阳光联动、MDRT、区域累计PVI、机构团队汇总、属地代理人和三类全年明细。"
)

if st.button("生成报告", type="primary", disabled=not (policy_file and agent_file)):
    if not cumulative_months:
        st.error("请至少选择一个累计统计月份。")
        st.stop()

    with st.spinner("正在清洗数据并生成 Excel 报告..."):
        with tempfile.TemporaryDirectory() as tmpdir:
            tmpdir_path = Path(tmpdir)
            policy_path = tmpdir_path / policy_file.name
            agent_path = tmpdir_path / agent_file.name
            output_path = tmpdir_path / "PVI经营报告.xlsx"
            policy_path.write_bytes(policy_file.getbuffer())
            agent_path.write_bytes(agent_file.getbuffer())

            # 更新全局配置，让 notebook 逻辑复用 Streamlit 页面选择。
            globals()["REPORT_MONTH"] = int(report_month)
            globals()["REPORT_MONTH_LABEL"] = f"{int(report_month)}月"
            globals()["CUMULATIVE_MONTHS"] = sorted(int(month) for month in cumulative_months)
            globals()["CUMULATIVE_MONTH_LABEL"] = "-".join(str(month) for month in CUMULATIVE_MONTHS) + "月"
            globals()["TRACKING_SHEET_NAME"] = f"{REPORT_MONTH_LABEL}安盛活力星达成追踪"
            globals()["REGION_MONTH_PVI_SHEET_NAME"] = f"区域{CUMULATIVE_MONTH_LABEL}PVI汇总"
            globals()["AGENCY_TEAM_PVI_SHEET_NAME"] = f"机构团队{REPORT_MONTH_LABEL}PVI汇总"
            globals()["AGENCY_TEAM_HIRES_SHEET_NAME"] = f"机构团队{REPORT_MONTH_LABEL}入职人数"
            globals()["CUMULATIVE_PVI_COLUMN"] = f"{CUMULATIVE_MONTH_LABEL}合计PVI"
            globals()["REPORT_MONTH_PVI_COLUMN"] = f"{REPORT_MONTH_LABEL}总PVI"
            globals()["REPORT_MONTH_HIRES_COLUMN"] = f"{REPORT_MONTH_LABEL}入职人数"
            globals()["MONTH_PVI_COLUMNS"] = [f"{month}月总PVI" for month in CUMULATIVE_MONTHS]
            globals()["MONTH_PVI_MAP"] = dict(zip(CUMULATIVE_MONTHS, MONTH_PVI_COLUMNS))
            globals()["SHEET_NAMES"] = {
                "tracking": TRACKING_SHEET_NAME,
                "sunshine": SUNSHINE_SHEET_NAME,
                "mdrt": "MDRT",
                "mdrt_region": "MDRT-区域",
                "local_agent": "大湾区属地代理人达成情况",
                "region_month_pvi": REGION_MONTH_PVI_SHEET_NAME,
                "agency_team_pvi": AGENCY_TEAM_PVI_SHEET_NAME,
                "agency_team_hires": AGENCY_TEAM_HIRES_SHEET_NAME,
                "kai_xuan_detail": "大湾区凯旋明细（全年）",
                "yaokun_detail": "曜坤区明细（全年）",
                "fangyuan_hq_detail": "方圆区总部直辖明细（全年）",
            }
            globals()["FINAL_SHEETS"] = list(SHEET_NAMES.values())
            globals()["SELECTED_AGENT_INFO_FILE"] = agent_path
            globals()["INPUT_FILE"] = policy_path
            globals()["DATA_FOLDER"] = policy_path.parent
            globals()["FILE_PATTERN"] = "保单汇总列表*.xlsx"

            requested_date = report_date_text.strip() or None
            df = clean_data(policy_path, SHEET_NAME, TEAM_KEYWORD)
            report_date = determine_report_date(df, requested_date)
            reports = calculate_reports(df=df, report_date=report_date, threshold=PVI_THRESHOLD)
            export_excel(reports, output_path, report_date)
            excel_bytes = output_path.read_bytes()

    st.success(f"报告已生成：{report_date:%Y-%m-%d}")

    agent_hire_summary = reports.get("agent_hire_summary", {})
    col1, col2, col3 = st.columns(3)
    col1.metric("今日新增", f"{agent_hire_summary.get('today_new_hires', 0)} 人")
    col2.metric("本月累计新增", f"{agent_hire_summary.get('month_new_hires', 0)} 人")
    col3.metric("输出 Sheet", f"{len(FINAL_SHEETS)} 个")

    st.download_button(
        "下载 Excel 报告",
        data=excel_bytes,
        file_name=f"PVI经营报告_{report_date:%Y%m%d}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.subheader("本次输出 Sheet")
    st.write(FINAL_SHEETS)
else:
    st.warning("请先上传保单汇总列表和代理人信息表。")
